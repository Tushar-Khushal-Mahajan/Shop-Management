'''*****IMPORT MODULES****'''
from tkinter import *
from functools import partial
from tkinter import ttk
from PIL import ImageTk,Image
from tkinter import messagebox
import mysql.connector
import re
import mysql.connector as connector
from tkcalendar import DateEntry
from datetime import date,datetime,timedelta
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import xlsxwriter
import openpyxl
'''
////////////////////////////////////////
\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
'''
#######################################################################################################################

class Accounting:
    def __init__(self,root):
            self.root = root

            self.root.state("zoomed")
            self.root.resizable(False,False)
            self.root.title("Accounting Window")

            bg  =Image.open(r"image\accounting_bg.png")
            bg = bg.resize((1535,840),Image.ANTIALIAS)
            self.bg = ImageTk.PhotoImage(bg)
            Label(self.root,image=self.bg,bg='light gray').pack(fill=BOTH)

    ##        ===============================================================================================
            conn = self.create_conn()
            cursor = conn.cursor()
            cursor.execute("SELECT sh_id FROM shop WHERE sh_status='yes'")
            self.shop_no = cursor.fetchone()
            conn.close()
            print("shop no is = ",self.shop_no[0])
    ##        =================================================================================================
    ##                            ***********VARIABLES**********
            self.td_sales = StringVar()
            self.td_orders = StringVar()
            self.td_creadits = StringVar()
            self.total_db = StringVar()

            self.ttl_sales = StringVar()
            self.ttl_orders = StringVar()
            self.ttl_creadits = StringVar()

            self.value_assigning()
    ##        =============================================================================================                    
            img1  =Image.open(r"image\back.png")
            img1 = img1.resize((50,30),Image.ANTIALIAS)
            self.backbg = ImageTk.PhotoImage(img1)
            Button(self.root, image=self.backbg,width=50,height=30,command=self.back).place(x=0,y=0)

                
            self.today_sales = LabelFrame(self.root, text="Total Sales Today", font=('Algerian',15))
            self.today_sales.place(x=20,y=50,width=270,height=110)
            Label(self.today_sales,textvariable=self.td_sales,text="30",font=('Times New Roman',15)).place(x=2,y=10)

            self.today_orders = LabelFrame(self.root, text="Total Orders Today", font=('Algerian',15))
            self.today_orders.place(x=350,y=50,width=270,height=110)
            Label(self.today_orders,textvariable=self.td_orders,text="2",font=('Times New Roman',15)).place(x=2,y=10)

            self.todayCustomer_creadits = LabelFrame(self.root, text="Todays Customer Creadits", font=('Algerian',12))
            self.todayCustomer_creadits.place(x=20,y=170,width=270,height=110)
            Label(self.todayCustomer_creadits,textvariable=self.total_db,text="total debtors :- 3",font=('Times New Roman',14)).place(x=2,y=10)
            Label(self.todayCustomer_creadits,textvariable=self.td_creadits,text="total amount :- 1000",font=('Times New Roman',14)).place(x=2,y=50)

    ##        ========================================================================================================================

            self.total_sales = LabelFrame(self.root, text="Total Sales", font=('Algerian',12))
            self.total_sales.place(x=20,y=330,width=270,height=110)
            Label(self.total_sales,textvariable=self.ttl_sales,text="Total Sales :- 33",font=('Times New Roman',14)).place(x=2,y=10)

            self.today_orders = LabelFrame(self.root, text="Total Orders", font=('Algerian',15))
            self.today_orders.place(x=350,y=330,width=270,height=110)
            Label(self.today_orders,textvariable=self.ttl_orders,text="Orders = 10",font=('Times New Roman',15)).place(x=2,y=10)

            self.total_creadits = LabelFrame(self.root, text="Total Customer Creadits", font=('Algerian',12))
            self.total_creadits.place(x=20,y=460,width=270,height=110)
            Label(self.total_creadits,textvariable=self.ttl_creadits,text="Total Creadits :- 3",font=('Times New Roman',14)).place(x=2,y=10)
##    ===========================================================================
##                Key Binding

            self.root.bind('<Control_L><b>',self.back)
    ##==============================================================================================================================================
##                ***************METHODS************

    def create_conn(self):
            conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
            return conn
        
    def value_assigning(self):
        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("select count(s_id) from confirm_selling where sh_id='{}'".format(self.shop_no[0]))#total sales
        total_sale = cursor.fetchone()

        cursor.execute("select  count(s_date) from confirm_selling where sh_id = '{}' and s_id in (select s_id from confirm_selling where s_date = date(sysdate())and sh_id='{}')".format(self.shop_no[0],self.shop_no[0]))
        today_sale = cursor.fetchone()

        cursor.execute("select count(o_id) from receive_orders where sh_id='{}'".format(self.shop_no[0]))
        total_order = cursor.fetchone()

        cursor.execute("select  count(o_id) from receive_orders where sh_id = '{}' and o_id in (select o_id from receive_orders where receive_date = date(sysdate())and sh_id='{}')".format(self.shop_no[0],self.shop_no[0]))
        today_order = cursor.fetchone()

        cursor.execute("select count(s_id) from creadit where sh_id='{}'".format(self.shop_no[0]))
        total_creadit = cursor.fetchone()

        cursor.execute("select  count(s_id) from creadit where sh_id = '{}' and s_id in (select s_id from creadit where date = date(sysdate())and sh_id='{}')".format(self.shop_no[0],self.shop_no[0]))
        today_creadit = cursor.fetchone()

        cursor.execute("select count(distinct mo_no) from creadit where sh_id='{}'".format(self.shop_no[0]))
        total_db = cursor.fetchone()
        conn.close()
        
        try: 
            self.td_sales.set("Todays sales = "+str(today_sale[0]))
            self.td_orders.set("Todays Orders = "+str(today_order[0]))
            self.td_creadits.set("Todays Creadits = "+str(today_creadit[0]))
            self.total_db.set("No Of Debtors = "+str(total_db[0]))

            self.ttl_sales.set("Total Sales = "+str(total_sale[0]))
            self.ttl_orders.set("Total Orders = "+str(total_order[0]))
            self.ttl_creadits.set("Total Creadits = "+str(total_creadit[0]))
        except Exception:
            pass

    def back(self,*args):
            self.root.withdraw()
            self.stock_w = Toplevel()
            obj = stock(self.stock_w)


#######################################################################################################################

class Order:
    def __init__(self,root):
        self.root = root

        self.root.state("zoomed")
        self.root.resizable(False,False)
        self.root.title("Order Window")
        #---------------------------------------------------------------------

        topFrame = LabelFrame(self.root,text="Fill The Details",font=("ALGERIAN",14),bd=10,relief=RIDGE)
        topFrame.place(x=0,y=3,width=1530,height=250)
        bg  =Image.open(r"image\bg3.png")
        bg = bg.resize((1505,217),Image.ANTIALIAS)
        self.bg = ImageTk.PhotoImage(bg)
        self.mainmid_frame = Label(topFrame,image=self.bg).place(x=0,y=0)



        autoFillFrame = LabelFrame(self.root,text="Auto Order",font=("ALGERIAN",10),bd=6,relief=RIDGE)
        autoFillFrame.place(x=900,y=20,width=620,height=225)
        bg1  =Image.open(r"image\bg3.png")
        bg1 = bg1.resize((607,226),Image.ANTIALIAS)
        self.bg1 = ImageTk.PhotoImage(bg1)
        self.side_frame = Label(autoFillFrame,image=self.bg1).place(x=0,y=0)

        treeFrame = Frame(self.root,bd=10,relief=RIDGE)
        treeFrame.place(x=0,y=255,width=900,height=580)

        sideFrame = Frame(self.root, bd=10,relief=RIDGE)
        sideFrame.place(x=900,y=255,width=630,height=580)
        bg12  =Image.open(r"image\bg3.png")
        bg12 = bg12.resize((607,555),Image.ANTIALIAS)
        self.bg12 = ImageTk.PhotoImage(bg12)
        self.side_frame = Label(sideFrame,image=self.bg12).place(x=0,y=0)

        fileFrame = Frame(sideFrame, bd=3, relief=RIDGE)
        fileFrame.place(x=10,y=10,width=590, height=470)
        Label(fileFrame, text="***********Order Receiving Frame***********",font=("Times New Roman",18)).pack()

        #----------------------------------------------------------------------
        '''============== variables ==========================================='''
        self.sh_id = StringVar()
        self.sh_name = StringVar()

        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT sh_id FROM shop WHERE sh_status='yes'")
        self.shop_no = cursor.fetchone()

        cursor.execute("SELECT mono FROM shop WHERE sh_status='yes'")
        self.mono = cursor.fetchone()

        cursor.execute("SELECT sh_name FROM shop WHERE sh_status='yes'")
        self.shop_name = cursor.fetchone()

        cursor.execute("select o_id from orders where sh_id = '{}' order by o_id desc".format(self.shop_no[0]))
        self.order_id = cursor.fetchone()
        if self.order_id == None: #if the order no is 
            self.order_id = 1
        else:
            self.order_id = self.order_id[0]
            self.order_id += 1
        print("order id = ",)
        conn.close()

        print(self.shop_no[0])
        #----------------------------------------------------------------------
        '''=========================top frame coding========================='''
        #variables--
        self.product_name = StringVar()
        self.order_QTY  = StringVar()
        self.e_oid = StringVar()

        '''----------------------------------------------------------------------'''
        img1  =Image.open(r"image\back.png")
        img1 = img1.resize((50,30),Image.ANTIALIAS)
        self.backbg = ImageTk.PhotoImage(img1)
        Button(topFrame, image=self.backbg,width=50,height=30,command=self.back).place(x=0,y=0)
        
        self.temp = Entry(topFrame).place(x=0,y=0,width=0,height=0)
        
        Label(topFrame,text="Product Name (F1)",bg='light gray',fg="black",font=("Time New Roman",14)).place(x=58,y=15)
        self.p_name = Entry(topFrame,textvariable = self.product_name,font=("Times New Roman",15),border=4,relief=RIDGE)
        self.p_name.place(x=20,y=45,width=230,height=30)

        Label(topFrame,text="Order QTY",bg='light gray' ,font=("Time New Roman",14)).place(x=420,y=15)
        self.O_Qty = Entry(topFrame,textvariable = self.order_QTY,font=("Times New Roman",15),border=4,relief=RIDGE)
        self.O_Qty.place(x=350,y=45,width=230,height=30)


        #Label(topFrame,text="Order QTY",bg='light gray' ,font=("Time New Roman",14)).place(x=420,y=15)
        
        self.o_id = Entry(topFrame,textvariable = self.e_oid,state='disabled',font=("Times New Roman",15,'bold'),border=4,relief=RIDGE)
        self.o_id.place(x=20,y=150,width=230,height=30)
        self.e_oid.set("order id = "+str(self.order_id))
        

        bg2  =Image.open(r"image\insert new.png")
        bg2 = bg2.resize((120,45),Image.ANTIALIAS)
        self.bg2 = ImageTk.PhotoImage(bg2)
        Button(topFrame, text="Insert (ctrl+i)",image=self.bg2,bg="black" ,command=self.insert_Btn).place(x=600,y=45,width=120,height=50)

        bg3  =Image.open(r"image\update new.png")
        bg3 = bg3.resize((120,45),Image.ANTIALIAS)
        self.bg3 = ImageTk.PhotoImage(bg3)
        Button(topFrame, text="Update (ctrl+u)",image=self.bg3,bg="black" ,command=self.update_Btn).place(x=750,y=45,width=120,height=50)

        bg4  =Image.open(r"image\remove new.png")
        bg4 = bg4.resize((120,45),Image.ANTIALIAS)
        self.bg4 = ImageTk.PhotoImage(bg4)
        Button(topFrame, text="Remove (ctrl+r)",image=self.bg4,bg="black", command=self.delete_Btn).place(x=600,y=110,width=120,height=50)

        bg5  =Image.open(r"image\clear new.png")
        bg5 = bg5.resize((120,45),Image.ANTIALIAS)
        self.bg5 = ImageTk.PhotoImage(bg5)
        Button(topFrame, text="Clear (ctrl+c)",image=self.bg5,bg="black", command=self.clear_Btn).place(x=750,y=110,width=120,height=50)

        bg13  =Image.open(r"image\search new.png")
        bg13 = bg13.resize((120,45),Image.ANTIALIAS)
        self.bg13 = ImageTk.PhotoImage(bg13)
        Button(topFrame, text="Search (ctrl+s)",image=self.bg13,bg="black", command=self.search_Btn).place(x=600,y=170,width=120,height=50)


        bg9  =Image.open(r"image\refresh new.png")
        bg9 = bg9.resize((120,45),Image.ANTIALIAS)
        self.bg9 = ImageTk.PhotoImage(bg9)
        Button(topFrame, text="Reseat (alt)",image=self.bg9,bg="black", command=self.refresh_Btn).place(x=750,y=170,width=120,height=50)


        self.p_name.focus()
        #----------------------------------------------------------------------
        '''=========================auto fill frame coding========================='''
        
        bg6  =Image.open(r"image\low stock insert.png")
        bg6 = bg6.resize((200,60),Image.ANTIALIAS)
        self.bg6 = ImageTk.PhotoImage(bg6)
        Button(autoFillFrame, text="Insert Low Stock (ctrl+l)",image=self.bg6, command=self.lowStock_Btn).place(x=80,y=30,width=200,height=60)

        bg7  =Image.open(r"image\mid stock insert.png")
        bg7 = bg7.resize((200,60),Image.ANTIALIAS)
        self.bg7 = ImageTk.PhotoImage(bg7)
        Button(autoFillFrame, text="Insert Mid Stock (ctrl+m)",image=self.bg7, command=self.midStock_Btn).place(x=320,y=30,width=200,height=60)

        bg8  =Image.open(r"image\remove options new.png")
        bg8 = bg8.resize((200,60),Image.ANTIALIAS)
        self.bg8 = ImageTk.PhotoImage(bg8)
        Button(autoFillFrame, text="Reseat Order(ctrl+shift+r)",image=self.bg8,command=self.removeOptions_Btn ).place(x=200,y=110,width=200,height=60)

        bg15  =Image.open(r"image\setting.jpg")
        bg15 = bg15.resize((50,40),Image.ANTIALIAS)
        self.bg15 = ImageTk.PhotoImage(bg15)
        Button(autoFillFrame,text="Setting",image=self.bg15 ,command=self.settingIcon).place(x=555, y=165, width=50, height=40)

        #----------------------------------------------------------------------
        '''=========================tree frame coding========================='''
        sc_x = ttk.Scrollbar(treeFrame, orient=HORIZONTAL)
        sc_x.pack(side=BOTTOM, fill=X)

        sc_y = ttk.Scrollbar(treeFrame, orient=VERTICAL)
        sc_y.pack(side=RIGHT, fill=Y)
        
        self.orderTree = ttk.Treeview(treeFrame,xscrollcommand=sc_x,yscrollcommand=sc_y)

        self.orderTree['columns'] = ("p_id","p_name","o_qty")

        sc_x.config(command=self.orderTree.xview)
        sc_y.config(command=self.orderTree.yview)

        self.orderTree.column("#0", stretch=NO,width=0)
        self.orderTree.column("p_id",minwidth=100,width=50,anchor=CENTER)
        self.orderTree.column("p_name",minwidth=400,width=500,anchor=CENTER)
        self.orderTree.column("o_qty",minwidth=120,width=120,anchor=CENTER)
       
        self.orderTree.bind("<ButtonRelease-1>",self.order_tree_focus)
        

        self.orderTree.heading("#0", text="")
        self.orderTree.heading("p_id", text="Product ID", anchor=CENTER)
        self.orderTree.heading("p_name", text="Product Name", anchor=CENTER)
        self.orderTree.heading("o_qty", text="Order QTY", anchor=CENTER)
        self.orderTree.pack(fill=BOTH,expand=True)

        #----------------------------------------------------------------------
        '''=========================side frame coding========================='''
        self.orderId = StringVar()
        self.orderId.set(self.order_id)
        

        Label(fileFrame, text="Enter Received",font=('Times New Roman',14)).place(x=15,y=40)
        Label(fileFrame, text="Order id *",font=('Times New Roman',14)).place(x=15,y=60)
        self.r_oId = Entry(fileFrame,font=('Times New Roman',12),bd=5)
        self.r_oId.place(x=150,y=45,height=30)

        Label(fileFrame, text="Receiving Date",font=('Times New Roman',14)).place(x=15,y=90)
        Label(fileFrame, text="(yy-mm-dd)",font=('Times New Roman',14)).place(x=15,y=110)
        self.r_date = Entry(fileFrame,font=('Times New Roman',12),bd=5)
        self.r_date.place(x=150,y=100,height=30)

        Label(fileFrame, text="Supplier",font=('Times New Roman',14)).place(x=15,y=150)
        self.supplyer = Entry(fileFrame,font=('Times New Roman',12),bd=5)
        self.supplyer.place(x=150,y=150,height=30)

        Label(fileFrame, text="Enter Exel Sheet Path *",font=('Times New Roman',15)).place(x=15,y=200)
        self.received_path = Text(fileFrame, font=("Times New Roman",15),bd=5)
        self.received_path.place(x=55,y=240,height=50,width=450)

        Label(fileFrame, text="Enter First row No :- ",font=('Times New Roman',12)).place(x=15,y=300)
        self.f_row = Entry(fileFrame,font=('Times New Roman',12),bd=5)
        self.f_row.place(x=180,y=300,height=30,width=80)
        
        Label(fileFrame, text="Enter Last row No :- ",font=('Times New Roman',12)).place(x=300,y=300)
        self.l_row = Entry(fileFrame,font=('Times New Roman',12),bd=5)
        self.l_row.place(x=460,y=300,height=30,width=80)
        
        Label(fileFrame, text="Enter First Column No :- ",font=('Times New Roman',12)).place(x=15,y=350)
        self.f_col = Entry(fileFrame,font=('Times New Roman',12),bd=5)
        self.f_col.place(x=180,y=350,height=30,width=80)
        
        Label(fileFrame, text="Enter Last Column No :- ",font=('Times New Roman',12)).place(x=300,y=350)
        self.l_col = Entry(fileFrame,font=('Times New Roman',12),bd=5)
        self.l_col.place(x=460,y=350,height=30,width=80)

        Button(fileFrame, text="Add products From Exel",command=self.addFrom_Exel,relief=RIDGE, bd=5, font=("algerian",16)).place(x=30,y=400,width=300,height=60)
        Button(fileFrame, text="Clear",relief=RIDGE,bd=5,command=self.clear_order,font=("algerian",14)).place(x=350,y=400,width=200,height=60)


        bg11  =Image.open(r"image\save recipt.png")
        bg11 = bg11.resize((270,60),Image.ANTIALIAS)
        self.bg11 = ImageTk.PhotoImage(bg11)
        Button(sideFrame, text="Save Recipt(ctrl+shift+s)",image=self.bg11 , command=self.save_Recipt).place(x=30,y=490,width=260,height=60)

        bg10  =Image.open(r"image\next order.png")
        bg10 = bg10.resize((260,60),Image.ANTIALIAS)
        self.bg10 = ImageTk.PhotoImage(bg10)
        Button(sideFrame,image=self.bg10,text="Next Order(ctrl+shift+N)").place(x=320,y=490,width=260,height=60)

        #--------------
        self.productmainframeOpen = False
        
        self.target = -1

        self.isReseatFrameOpen = False

        self.isSettingFrameOpen = False

        #==============BINDING============================================================
        self.root.bind('<Control_L><b>',self.back)
        self.product_name.trace('w',self.pro_nameEnter)
        self.p_name.bind('<Down>',self.p_nameDown)
        self.p_name.bind('<Up>',self.p_nameUp)
        self.p_name.bind('<Return>',self.p_nameEnter)
        self.root.bind('<F1>',lambda event="":self.p_name.focus())
        self.root.bind('<Control_L><i>',self.insert_Btn)
        self.root.bind('<Control_L><u>',self.update_Btn)
        self.root.bind('<Control_L><r>',self.delete_Btn)
        self.root.bind('<Control_L><c>',self.clear_Btn)
        self.O_Qty.bind('<Return>',lambda events="": self.insert_Btn())
        self.root.bind('<Control_L><s>',self.search_Btn)
        self.root.bind('<F5>',self.refresh_Btn)

        self.root.bind('<Control_L><l>',self.lowStock_Btn)
        self.root.bind('<Control_L><m>',self.midStock_Btn)
        self.p_name.bind('<Control-Shift-R>',self.removeOptions_Btn)

        self.root.bind('<Control-Shift-S>',self.save_Recipt)

        self.root.bind('<Control-i>',self.settingIcon)
        self.root.bind('<Control_L><i>',self.settingIcon)

        self.m_list = []
        self.m_list_id = []


    #-------------------VARIABLES FOR AUTO ORDER WINDOW==========================================
        conn = self.createS_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT l_stock_qty,m_stock_qty,saving_path FROM order_ws WHERE sh_id='{}'".format(self.shop_no[0]))
        self.outo_OrderQty = cursor.fetchone()
        conn.close()

        #self.outo_OrderQty = 10


        self.prod_Id = 0
        
    #------------------------------METHOD SECTION---------------------------------------------------------------------------------------
    def create_conn(self):
                conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                return conn

    def createS_conn(self):
                conn = mysql.connector.connect(host='localhost', database='datasetting', user='root', password='Tushar()mysql[123]')
                return conn
        
    def back(self,*args):
            self.root.withdraw()
            self.stock_w = Toplevel()
            obj = stock(self.stock_w)
            
            
    def insert_Btn(self,*args):
        if  self.p_name.get().strip()=="" or self.O_Qty.get().strip()=="" or self.O_Qty.get() == '0' or self.O_Qty.get().isnumeric() == False:
            messagebox.showinfo("Info","Plz fill correct input")

        else:                
                p_name = self.p_name.get()
                o_qty = self.O_Qty.get()
                p_id = self.prod_Id

                try:
                            
                        conn = self.create_conn()
                        cursor = conn.cursor()
                        cursor.execute("SELECT *FROM products WHERE sh_id='{}' AND p_name='{}'".format(self.shop_no[0],p_name))
                        is_avl = cursor.fetchone()

                        print("product is available = ",is_avl)

                        if is_avl == None: #check whether the product is new product or exesting product..
                            ch = messagebox.askquestion("Info","This is new product do you want to order them.")
                        
                            if ch=='yes':
                                    p_id=0

                                    print(self.shop_no[0],p_id,self.order_id, p_name, o_qty)

                                    conn = self.create_conn()
                                    cursor = conn.cursor()
                                    
                                    cursor.execute("INSERT INTO orders(sh_id, p_id, o_id, p_name, o_qty) VALUES('{}','{}','{}','{}','{}')".format(self.shop_no[0],p_id,self.order_id,p_name,o_qty))
                                    conn.commit()

                                    print("p_id for existiong product = ",p_id)
                                    conn.close()

                        else:
                            conn = self.create_conn()
                            cursor = conn.cursor()
                            cursor.execute("INSERT INTO orders(sh_id, p_id, o_id, p_name, o_qty) VALUES('{}','{}','{}','{}','{}')".format(self.shop_no[0],p_id,self.order_id,p_name,o_qty))
                            conn.commit()
                            conn.close()

                        self.refresh_orderTree()
                
                except Exception:
                        messagebox.showerror("Error","Product already exist")
                self.clear_Btn()
                

    def update_Btn(self,*args):
        if  self.p_name.get().strip()=="" or self.O_Qty.get().strip()=="" or self.O_Qty.get() == '0' or self.O_Qty.get().isnumeric() == False:
                messagebox.showinfo("Info","Plz fill correct input")

        else:
            a = messagebox.askquestion("Que","Do you want to Update this record..")
            
            if a == 'yes':
                p_name = self.p_name.get()
                o_qty = self.O_Qty.get()

                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("UPDATE orders SET o_qty = '{}', p_name='{}' WHERE p_name='{}'AND sh_id = '{}' AND o_id='{}' ".format(o_qty,p_name,p_name, self.shop_no[0], self.order_id))
                conn.commit()
                conn.close()
                messagebox.showinfo("Info","Record Update successfully..")

                self.refresh_orderTree()


    def delete_Btn(self,*args):
        if  self.p_name.get().strip()=="" or self.O_Qty.get().strip()=="":
                messagebox.showinfo("Info","Plz fill correct input")

        else:
            a = messagebox.askquestion("Que","Do you want to Delete this record..")
            
            if a == 'yes':
                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM orders WHERE p_name = '{}' AND o_id = '{}'  AND sh_id = '{}'".format(self.p_name.get(), self.order_id, self.shop_no[0]))
                conn.commit()
                conn.close()

            self.refresh_orderTree()
            self.clear_Btn()

    def clear_Btn(self,*args):
        self.p_name.delete(0,END)
        self.O_Qty.delete(0,END)
        self.p_name.focus()


    def search_Btn(self,*args):
        if self.p_name.get().strip()=="":
            messagebox.showinfo("Info","plz select product first")
        else:

            p_name = "%"+self.p_name.get()+"%"
            conn = self.create_conn()
            cursor = conn.cursor()
            cursor.execute("SELECT p_id,p_name,o_qty FROM orders WHERE p_name like '{}' AND sh_id='{}' AND o_id='{}'".format(p_name, self.shop_no[0], self.order_id))
            data = cursor.fetchall()
            conn.close()

            self.orderTree.delete(*self.orderTree.get_children()) #clear treeview
            count = 0
            for i in data:
                    self.orderTree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2]))
                    count = count + 1
    

    def refresh_Btn(self,*args):
        self.clear_Btn()
        self.refresh_orderTree()        

    def refresh_orderTree(self):
        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("select *from orders where sh_id='{}' and o_id='{}'".format(self.shop_no[0],self.order_id))
        data = cursor.fetchall()
        conn.close()
        
        self.orderTree.delete(*self.orderTree.get_children()) #clear treeview
        count = 0
        for i in data:
                self.orderTree.insert(parent="", index="end", iid=count, text="" , values=(i[1],i[3],i[4]))
                count = count + 1


    def order_tree_focus(self,*args):
        try:
            self.clear_Btn()

            cursor_row = self.orderTree.focus()
            content = self.orderTree.item(cursor_row)
            focused_row = content['values']
            print(focused_row)
            self.product_name.set(focused_row[1])
            self.order_QTY.set(focused_row[2])
            self.p_name.focus()
        except Exception:
            pass

    #===AUTO ORDER FRAME CODING----------------------------------------------
        
    def orderTreeLowAndHighStock(self,qty): #this function enter low and high stock in treeview 
        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("select p_id,p_name from orders where sh_id='{}' and o_id='{}'".format(self.shop_no[0],self.order_id))
        data = cursor.fetchall()
        conn.close()
        
        self.orderTree.delete(*self.orderTree.get_children()) #clear treeview
        count = 0
        for i in data:
                self.orderTree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],qty))
                count = count + 1
        self.refresh_orderTree()

    def autoInsertLowAndHighStock(self,data,qty): #this function insert a low and high stock in a orders table in db
         try:
             conn = self.create_conn()
             cursor = conn.cursor()

             for a in data:
                 cursor.execute("INSERT INTO orders VALUES('{}','{}','{}','{}','{}')".format(self.shop_no[0],a[0],self.order_id, a[1], qty))
                 conn.commit()
             conn.close()
         
             self.refresh_orderTree()

         except Exception:
                messagebox.showerror("Error","Product already exist")
         self.clear_Btn()
         
                
    def lowStock_Btn(self,*args):  #this function fetches a low stock after clicking a button
        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT p_id, p_name FROM products where sh_id='{}' and p_id not in(select p_id from orders where sh_id='{}' AND o_id='{}') and p_qty<='{}' order by p_qty asc".format(self.shop_no[0],self.shop_no[0],self.order_id,10))
        data = cursor.fetchall()
        conn.close()



        print("low stock = ",data)
        self.autoInsertLowAndHighStock(data,self.outo_OrderQty[0])
        self.orderTreeLowAndHighStock(self.outo_OrderQty[0])
    

    def midStock_Btn(self,*args):   #this function fetches a mid stock after clicking a button
        conn = self.create_conn()
        cursor = conn.cursor()
      
        cursor.execute("SELECT p_id,p_name FROM products where sh_id='{}' and p_id not in(select p_id from orders where sh_id='{}' AND o_id='{}') and p_qty>='{}' AND p_qty<='{}' order by p_qty asc".format(self.shop_no[0],self.shop_no[0],self.order_id,11,20))
        data = cursor.fetchall()
        conn.close()

        print("mid stock = ",data)
        self.autoInsertLowAndHighStock(data,self.outo_OrderQty[1])
        self.orderTreeLowAndHighStock(self.outo_OrderQty[1])
    
    #---------

    def removeOptions_Btn(self,*args):
        def reseatLowStock(*args):
            self.clear_Btn()
            ch = messagebox.askquestion("Que window","Do you want to remove low stock")

            if ch == 'yes':
                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM orders WHERE sh_id='{}' AND o_id='{}' AND p_id in (SELECT p_id FROM products where sh_id='{}' AND p_qty<='{}')".format(self.shop_no[0],self.order_id,self.shop_no[0],10))
                conn.commit()
                conn.close()
                messagebox.showinfo("info","low stock removed")
                self.refresh_orderTree()
                
                try:
                    self.reseat_Frame.destroy()
                except Exception:
                    pass

        def reseatMidStock(*args):
            self.clear_Btn()
            ch = messagebox.askquestion("Que window","Do you want to remove mid stock")

            if ch == 'yes':
                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM orders WHERE sh_id='{}' AND o_id='{}' AND p_id in (SELECT p_id FROM products where sh_id='{}' AND p_qty>='{}' AND p_qty<='{}')".format(self.shop_no[0],self.order_id,self.shop_no[0],11,20))
                conn.commit()
                conn.close()
                messagebox.showinfo("info","mid stock removed")
                self.refresh_orderTree()
                
                try:
                    self.reseat_Frame.destroy()
                except Exception:
                    pass

        def allReseat(*args):
            self.clear_Btn()
            ch = messagebox.askquestion("Que window","Do you want to remove all")

            if ch == 'yes':
                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM orders WHERE sh_id='{}' AND o_id='{}' AND p_id in (SELECT p_id FROM products where sh_id='{}')".format(self.shop_no[0],self.order_id,self.shop_no[0]))
                conn.commit()
                conn.close()
                messagebox.showinfo("info","reseted.")
                self.refresh_orderTree()
                
                try:
                    self.reseat_Frame.destroy()
                except Exception:
                    pass

        def binding():
            self.root.bind('<Control-Alt_L><l>',reseatLowStock)
            self.root.bind('<Control-Alt_L><m>',reseatMidStock)
            self.root.bind('<Control-Alt_L><a>',allReseat)


        #---------------------------------------
        if self.isReseatFrameOpen == False:
            self.reseat_Frame = LabelFrame(self.root,text="SELECT RESEAT OPTION", relief=RIDGE,bd=5 , fg="red")
            self.reseat_Frame.place(x=970,y=270,width=500,height=250)
            #print("reseat frame open= ",self.isReseatFrameOpen)
            Button(self.reseat_Frame, text="Reseat Low Stock", command=reseatLowStock).place(x=20,y=10,width=200,height=60)
            Button(self.reseat_Frame, text="Reseat Mid Stock", command=reseatMidStock).place(x=265,y=10,width=200,height=60)
            Button(self.reseat_Frame, text="Reseat All", command=allReseat).place(x=150,y=90,width=200,height=60)

            Label(self.reseat_Frame, text="For closing this window click Reseat Order button again.").pack(side=BOTTOM)
            self.isReseatFrameOpen=True

            '''binding'''
            binding()
            

            #print("is reseat frame open = ",self.isReseatFrameOpen)
        else:
            self.isReseatFrameOpen=False
            self.reseat_Frame.destroy()



    #==============text widget Btn coding==========================================
    def clear_order(self,*args):
        self.l_col.delete(0,END)
        self.f_col.delete(0,END)
        self.l_row.delete(0,END)
        self.f_row.delete(0,END)
        self.received_path.delete(1.0,END)
        self.supplyer.delete(0,END)
        self.r_date.delete(0,END)
        self.r_oId.delete(0,END)

    def readFrom_Exel(self,*args):
            try:
                path = str(self.received_path.get(1.0,END).strip())
                workbook = openpyxl.load_workbook(r""+path)
                sheets = workbook.sheetnames
                print(sheets)
            except Exception:
                messagebox.showerror("Note: ","Provide valid path")
                
            try:
                srow = int(self.f_row.get())
                scolumn = int(self.f_col.get())
                lrow = int(self.l_row.get())
                lcolumn = int(self.l_col.get())
            except Exception as e:
                messagebox.showerror("Note: ","Provide valid integer column or row no..")

            try:
                sheet1 = workbook[sheets[0]]

                mainData = []
                for i in range(srow,lrow+1):
                    s=[]
                    for j in range(scolumn,lcolumn+1):
                        #print(str(sheet1.cell(i,j).value)+"\t\t", end="")
                        s.append(sheet1.cell(i,j).value)
                        #print(s)
                    mainData.append(s)
                ##    print(mainData)
                workbook.close()
                return mainData
            
            except Exception as e:
                print(e)
            


    #responsible to add products into exel file..
    def addFrom_Exel(self,*args):
        if not(self.l_col.get().strip()=="" or self.f_col.get().strip()=="" or self.l_row.get().strip()=="" or self.f_row.get().strip()=="" or self.received_path.get(1.0,END).strip()=="" or self.supplyer.get().strip()=="" or self.r_date.get().strip()=="" or self.r_oId.get().strip()==""):
                    ch = messagebox.askquestion("Quistion","Do you want to add products from Exel sheet")

                    if ch=='yes':
                        mainData = self.readFrom_Exel()
                        print("data from exel = ",mainData)

                        conn = self.create_conn()
                        cursor = conn.cursor()
                        cursor.execute("SELECT p_id FROM products WHERE sh_id={} order by p_id desc".format(self.shop_no[0]))
                        id = cursor.fetchone()
                        conn.close()

                        if id==None:
                            id = 101
                        else:
                            id = id[0]+1
                            print("latest id is = ",id)
                        try:
                                for a in mainData:
                                    if a[1] == 0: #for new products where product id is {0}
                                        print("new = ",a)
                                        conn = self.create_conn()
                                        cursor = conn.cursor()
                                        cursor.execute("INSERT INTO products(sh_id,p_id,p_name,p_qty,p_price,s_price) VALUES('{}','{}','{}','{}','{}','{}')".format(self.shop_no[0],a[1],a[2],a[3],a[4],a[5]))
                                        conn.commit()
                                        conn.close()
                                        
                                        
                                    else:      #for old products where product is non {0}
                                        print("old = ",a)
                                        conn = self.create_conn()
                                        cursor = conn.cursor()
                                        if a[4] == None:
                                            print("inside if block")
                                            cursor.execute("update products set p_qty= p_qty+{} where p_id={} and sh_id={}".format(int(a[3]),a[1],self.shop_no[0]))
                                            conn.commit()
                                        else:
                                            print("inside else block")
                                            print("p id",a[1])
                                            print("p qty",int(a[3]))
                                            print("p price",a[4])
                                            print("s price",a[5])
                                            print(self.shop_no[0])
                                            cursor.execute("update products set p_qty= p_qty+{},p_price='{}',s_price='{}' where p_id={} and sh_id={}".format(int(a[3]),a[4],a[5],a[1],self.shop_no[0]))
                                            conn.commit()
                                            
                                        conn.close()
                                        
                                conn = self.create_conn()
                                cursor = conn.cursor()
                                cursor.execute("INSERT INTO receive_orders(sh_id,o_id,supplier,receive_date) VALUES('{}','{}','{}','{}')".format(self.shop_no[0],self.r_oId.get(),self.supplyer.get(),self.r_date.get()))
                                conn.commit()
                                conn.close()
                                messagebox.showinfo("Info: ","Added Successfully..")
                                self.clear_order()
                                
                        except Exception as e:
                            print("Exception ",e)

        else:
            messagebox.showerror("Note","All fields are required")
    #==============================================================================
    def save_Recipt(self,*args):

        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT p_id,p_name, o_qty FROM orders WHERE sh_id='{}' AND o_id='{}'".format(self.shop_no[0],self.order_id))
        data = cursor.fetchall()
        conn.close()

        print("data = ",data)
        
        
        ch = messagebox.askquestion("Question","Do you want to save..")

        if ch=='yes':

            try:
                print("ricipt is = \n",data)
                try:
                    path = r""+str(self.outo_OrderQty[2]).strip()+str(self.shop_name[0])+"order no = "+str(self.order_id)+".xlsx"
                except Exception:
##                    print("path is = ",self.outo_OrderQty[2])
                    path = "C:/shop management/order_recipts/"+str(self.shop_name[0])+" order no = "+str(self.order_id)+".xlsx"
                    messagebox.showinfo("info","your path is invalid your file temporary saved in location{C:/shop management/order_recipts/}")

                workbook = xlsxwriter.Workbook(path)
                worksheet = workbook.add_worksheet("Order Recipt")

                worksheet.write(0,0,'Order ID :- '+str(self.order_id))
                worksheet.write(1,0,'Shop Name :- '+self.shop_name[0])
                worksheet.write(2,0,'MO NO :- '+self.mono[0])
                
                row = 5
                worksheet.write(5,0,'#')
                worksheet.write(5,1,'Product ID')
                worksheet.write(5,2,'Product Name')
                worksheet.write(5,3,'Order QTY')
                row+=1

                for index,entry in enumerate(data):
                        worksheet.write(row,0,str(index+1))
                        worksheet.write(row,1,entry[0])
                        worksheet.write(row,2,entry[1])
                        worksheet.write(row,3,entry[2])
                        row+=1

                workbook.close()

                
                messagebox.showinfo("Info","File save successfully..")
            except Exception as e:
                #messagebox.showerror("message","something went wrong plz check provided path.")
                print(e)
    
    #----------- BELOW CODE WRITE FOR HIDING WINDOW    
    def pro_nameEnter(self,*args):
        if self.p_name.get().strip()== "":
                try:
                    self.productmainframeOpen = False
                    self.target=-1
                    self.productmainframe.destroy()
                except Exception:
                    pass
            
        elif self.productmainframeOpen == False:
                def tree_clicked(*args):
                    try:
                        cursor_row = self.sold_tree.focus()
                        content = self.sold_tree.item(cursor_row)
                        focused_row = content['values']
                        print(focused_row)
                        self.product_name.set(focused_row[1])
                        self.prod_Id = focused_row[0]
                                                
                        self.productmainframeOpen = False
                        self.target=-1
                        self.productmainframe.destroy()
                        self.p_name.focus()
                    except:
                        pass
            
                self.productmainframe = LabelFrame(self.root,text="select product",font=('Algerian',12,'bold'),bd=10,relief=RIDGE)
                self.productmainframe.place(x=100,y=310,width=800,height=400)

                self.productFrame = Frame(self.productmainframe,bd=5,relief=RIDGE)
                self.productFrame.place(x=3,y=35,width=780,height=335)

                #-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-
                self.count_down_cursor_value = 0
                #-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-
                
                img  =Image.open(r"image\exit.png")
                img = img.resize((90,35),Image.ANTIALIAS)
                self.exit = ImageTk.PhotoImage(img)
                Button(self.productmainframe, image=self.exit,command=self.exitbtn).place(x=690,y=0)


                sc_x = ttk.Scrollbar(self.productFrame, orient=HORIZONTAL)
                sc_x.pack(side=BOTTOM, fill=X)

                sc_y = ttk.Scrollbar(self.productFrame, orient=VERTICAL)
                sc_y.pack(side=RIGHT, fill=Y)

                self.sold_tree = ttk.Treeview(self.productFrame,xscrollcommand=sc_x,yscrollcommand=sc_y)
                self.sold_tree['columns'] = ("p_id","p_name","p_qty","p_price")

                self.sold_tree.column("#0",stretch=NO,width=0)
                self.sold_tree.column("p_id",minwidth=100,width=100,anchor=CENTER)
                self.sold_tree.column("p_name",minwidth=400,anchor=CENTER)
                self.sold_tree.column("p_qty",minwidth=100,anchor=CENTER)
                self.sold_tree.column("p_price",minwidth=150,anchor=CENTER)

                self.sold_tree.heading("#0",text="")
                self.sold_tree.heading("p_id",text="product id",anchor=CENTER)
                self.sold_tree.heading("p_name",text="product name",anchor=CENTER)
                self.sold_tree.heading("p_qty",text="QTY",anchor=CENTER)
                self.sold_tree.heading("p_price",text="price",anchor=CENTER)
                self.sold_tree.pack(fill=BOTH, expand=True)
                self.sold_tree.bind("<ButtonRelease-1>",tree_clicked)

                
                #self.sold_tree_insert(-1)
                self.sold_tree_insert()

                self.productmainframeOpen = True
            
        else:
                #self.sold_tree_insert(0)
                self.sold_tree_insert()

    def sold_tree_insert(self):
        
            proName = '%'+self.p_name.get()+'%'

            #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
            conn = self.create_conn()
            cursor=conn.cursor()

            cursor.execute("SELECT p_id,p_name,p_qty,s_price FROM products WHERE p_name like '{}' AND sh_id = '{}'  ".format(proName,self.shop_no[0]))
            #cursor.execute(''' select p_id,p_name,p_qty,s_price from products as a where sh_id = '{}' OR p_name like '{}' 
             #                   union select p_id,p_name,"","" from orders as b where b.p_id not in(select p_id from products where sh_id='{}') AND b.o_id='{}' AND b.sh_id='{}' OR b.p_name like '{}' '''.format(self.shop_no[0],proName,self.shop_no[0],self.order_id,self.shop_no[0],proName))
            self.soldTree_data = cursor.fetchall()

            cursor.execute(" SELECT count(p_name)as count FROM products WHERE p_name like '{}' AND sh_id = '{}' ".format(proName,self.shop_no[0]))
            c= cursor.fetchone()
            self.count_sell_treeviewValues = c[0]        
            conn.close()

            #print(self.count_sell_treeviewValues)
            
            self.sold_tree.delete(*self.sold_tree.get_children())  #clr treeview

            self.sold_tree.tag_configure('focus',background="blue",foreground="white", font=('Times New Roman',12,'bold'))
            self.sold_tree_insert_refresh()

    def sold_tree_insert_refresh(self):
            count = 0
            for i in self.soldTree_data:
               
                    if self.target == count:
                            self.sold_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3]),tags=('focus',))

                            self.content =self.sold_tree.item(self.target)
                            
                            self.highlightrow = self.content['values']
                            
                    else:
                            self.sold_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3]))
                    count = count + 1

    def p_nameUp(self,*args):
            if self.p_name.get() == "":
                    pass
            
            elif self.target != 0 and self.target != -1:
                    self.target = self.target-1
                    self.sold_tree_insert()
            
            else:
                    self.target = self.count_sell_treeviewValues-1
                    self.sold_tree_insert()
            
    
    def p_nameDown(self,event=""):
            if self.p_name.get() == "":
                    pass

            elif self.target < self.count_sell_treeviewValues-1:
                    self.target = self.target+1
                    self.sold_tree_insert()
            else:
                    self.target = 0
                    self.sold_tree_insert()

    def p_nameEnter(self,*args):
            if self.p_name.get()=="" or self.highlightrow == None or self.target == -1:
                    pass

            else:
                    
                    '''data fiill in the text boxes'''
                    self.prod_Id = self.highlightrow[0]
                    p_name = self.highlightrow[1]                        
                    self.product_name.set(p_name)
                    self.exitbtn()

    

    def exitbtn(self):
            self.productmainframeOpen = False
            self.target=-1 #new line add
            self.productmainframe.destroy()

    #-----------------------------------------------------------------------------------------------
    #--------------SETTING ICON---------------------------------------------------------------------
    
    def settingIcon(self,*args):
        self.lowStockSize = StringVar()
        self.midStockSize = StringVar()
                
        def close():
            self.setting_W.destroy()
            self.isSettingFrameOpen=False

        def save():
            if self.e_LowOrderSize.get().strip()=="" or self.e_MidOrderSize.get().strip()=="" or self.sheet_path.get(1.0,END).strip()=="":
                messagebox.showerror("message","Provide valid details..")
            else:
                try:
                    sconn = self.createS_conn()
                    cursor = sconn.cursor()
                    cursor.execute("UPDATE order_ws SET l_stock_qty={}, m_stock_qty={}, saving_path='{}' WHERE sh_id={}".format(self.e_LowOrderSize.get(),self.e_MidOrderSize.get(),self.sheet_path.get(1.0,END),self.shop_no[0]))
                    sconn.commit()
                    sconn.close()

                    conn = self.createS_conn()
                    cursor = conn.cursor()
                    cursor.execute("SELECT l_stock_qty,m_stock_qty,saving_path FROM order_ws WHERE sh_id='{}'".format(self.shop_no[0]))
                    self.outo_OrderQty = cursor.fetchone()
                    conn.close()

                    messagebox.showinfo("Success","Changes Apply Successfully..")
                    close()

                except Exception as e:
                    #messagebox.showerror("message","something went wrong plz check provided details..")
                  print(e)
        #-------------------------------------------------
        if self.isSettingFrameOpen == True:
                self.setting_W.destroy()
                self.isSettingFrameOpen=False
                
        else:
                self.setting_W = Frame(self.root,bg="#FFFFFF",relief=RIDGE,bd=5)
                self.setting_W.place(x=300,y=200,width=630,height=400)
                self.isSettingFrameOpen=True

                Button(self.setting_W ,text="Save",bd=5, command=save).place(x=350,y=340,width=120,height=50)
                Button(self.setting_W, text="close",bd=5 ,command=close).place(x=487,y=340,width=120,height=50)

                self.main_frame = Frame(self.setting_W,relief=RIDGE,bd=5)
                self.main_frame.place(x=10,y=3,width=600,height=330)

                topFrame = LabelFrame(self.main_frame,text="Auto Qty Size",relief=RIDGE,bd=5,width=567,height=200)
                topFrame.grid(row=1,column=0,ipadx=60)

                Label(topFrame,text="Low stock order size = ").grid(row=1,column=1)
                self.e_LowOrderSize = Entry(topFrame,textvariable=self.lowStockSize)
                self.e_LowOrderSize.grid(row=1,column=2,pady=5)
                
                Label(topFrame,text="Mid stock order size = ").grid(row=2,column=1)
                self.e_MidOrderSize = Entry(topFrame,textvariable=self.midStockSize)
                self.e_MidOrderSize.grid(row=2,column=2,pady=5)

                Label(self.main_frame,text="Enter valid path where you want to save your order exel sheet..").grid(row=2,column=0, pady=20)
                self.sheet_path = Text(self.main_frame,bd=5,width=68,height=5)
                self.sheet_path.grid(row=3,column=0,padx=5)

                self.lowStockSize.set(self.outo_OrderQty[0])
                self.midStockSize.set(self.outo_OrderQty[1])
                self.sheet_path.insert(1.0,self.outo_OrderQty[2])



#######################################################################################################################

class Creadit:
        def __init__(self,root,sell_id ,total_price,remaining_price):
                self.root = root
                self.root.state("zoomed")
                self.root.title("Creadit Page")
                self.root.resizable(False,False)


                '''----------------------- shop name and shop no ----------------------------'''
                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                conn = self.create_conn()
                cursor = conn.cursor()
                
                cursor.execute("SELECT sh_id FROM shop WHERE sh_status='yes' ")
                self.shop_no = cursor.fetchone()
                #cursor.execute("SELECT sh_name FROM shop WHERE sh_id={} ".format(self.shop_no[0]))
                #self.shop_name = cursor.fetchone()
                
                conn.close()
                
                print("shop no = "+str(self.shop_no[0]))

                '''-------------------------------FRAMES---------------------------------------'''
                
                cTopFrame = Frame(self.root,bd=10,relief=RIDGE)
                cTopFrame.place(x=0,y=0,width=1530,height=250)

                cMainFrame = Frame(self.root,bd=10, relief=RIDGE)
                cMainFrame.place(x=0,y=250,width=1530,height=570)
                cMainFrame.focus()
                

                treeFrame = Frame(cMainFrame)
                treeFrame.pack(fill=BOTH,expand=True)

                img1  =Image.open(r"image\counter_image6.jpg")
                img1 = img1.resize((1530,226),Image.ANTIALIAS)
                self.photo = ImageTk.PhotoImage(img1)
                self.mainmid_frame = Label(cTopFrame,image=self.photo).place(x=0,y=0)

                '''================================ top frame label and entries ======================'''
                #                               variables
                self.name = StringVar()
                self.mono = StringVar()
                self.total = StringVar()
                self.remaining = StringVar()
                self.sell_id = StringVar()
                self.cit = StringVar()
                self.rem = StringVar()


                #-------------------------------------------------------------------------------------
                
                img1  =Image.open(r"image\back.png")
                img1 = img1.resize((50,30),Image.ANTIALIAS)
                self.backbg = ImageTk.PhotoImage(img1)
                Button(cTopFrame, image=self.backbg,width=50,height=30,command=self.back).place(x=0,y=0)

                Label(cTopFrame, text="customer name :",font=('Sitka Small Semibold',14),bg="#87CEFA",fg="#000000").place(x=160,y=7)
                self.c_name = Entry(cTopFrame,textvariable=self.name,font=('Times New Roman',15),bd=5,relief=RIDGE)
                self.c_name.place(x=100,y=40,width=270,height=30)
                
                Label(cTopFrame, text="mobile no :",font=('Sitka Small Semibold',14),bg="#87CEFA",fg="#000000").place(x=515,y=7)
                self.c_mono = Entry(cTopFrame, textvariable=self.mono,font=('Times New Roman',14),bd=5,relief=RIDGE)
                self.c_mono.place(x=430,y=40,width=270)
                
                
                Label(cTopFrame, text="total price :",font=('Sitka Small Semibold',14),bg="#87CEFA",fg="#000000").place(x=830,y=7)
                self.total_price = Entry(cTopFrame, textvariable=self.total, font=('Times New Roman',14),bd=5,relief=RIDGE)
                self.total_price.place(x=750,y=40,width=270)
                
                Label(cTopFrame, text="remaining rs :",font=('Sitka Small Semibold',14),bg="#87CEFA",fg="#000000").place(x=1140,y=7)
                self.remaining_price = Entry(cTopFrame,textvariable=self.remaining , font=('Times New Roman',14),bd=5,relief=RIDGE)
                self.remaining_price.place(x=1080,y=40,width=260)

                Label(cTopFrame, text="selling id :",font=('Sitka Small Semibold',14),bg="#87CEFA",fg="#000000").place(x=100,y=94)
                self.selling_id = Entry(cTopFrame, textvariable=self.sell_id, font=('Times New Roman',14),bd=5,relief=RIDGE)
                self.selling_id.place(x=230,y=95,width=130)

                Label(cTopFrame, text="city :",font=('Sitka Small Semibold',14),bg="#87CEFA",fg="#000000").place(x=400,y=94)
                self.city = Entry(cTopFrame,textvariable=self.cit, font=('Times New Roman',14),bd=5,relief=RIDGE)
                self.city.place(x=470,y=95,width=130)

                img1  =Image.open(r"image\add.jpg")
                img1 = img1.resize((150,50),Image.ANTIALIAS)
                self.addbtn = ImageTk.PhotoImage(img1)
                Button(cTopFrame,text="insert",command=self.insertBtn,image=self.addbtn,bg="#87CEFA",font=('ALGERIAN',14,'bold')).place(x=850,y=100,width=150,height=50)

                img2  =Image.open(r"image\update.png")
                img2 = img2.resize((150,50),Image.ANTIALIAS)
                self.updatebtn = ImageTk.PhotoImage(img2)
                Button(cTopFrame,text="update",image=self.updatebtn,command=self.update,bg="#87CEFA",font=('ALGERIAN',14,'bold')).place(x=1050,y=100,width=150,height=50)                

                img3  =Image.open(r"image\delete.png")
                img3 = img3.resize((150,50),Image.ANTIALIAS)
                self.deletebtn = ImageTk.PhotoImage(img3)
                Button(cTopFrame,text="delete",image=self.deletebtn,command=self.delete,bg="#87CEFA",font=('ALGERIAN',14,'bold')).place(x=1250,y=100,width=150,height=50)

                img4  =Image.open(r"image\reseat.png")
                img4 = img4.resize((150,50),Image.ANTIALIAS)
                self.reseatBtn = ImageTk.PhotoImage(img4)
                Button(cTopFrame,text="reseat",image=self.reseatBtn,command=lambda *args: self.refresh_treeview(),bg="#87CEFA",font=('ALGERIAN',14,'bold')).place(x=1050,y=170,width=150,height=50)                
                
                img5  =Image.open(r"image\clear.png")
                img5 = img5.resize((150,50),Image.ANTIALIAS)
                self.clrbtn = ImageTk.PhotoImage(img5)
                Button(cTopFrame,text="clr",image=self.clrbtn,command=self.clearBtn,bg="#87CEFA",font=('ALGERIAN',14,'bold')).place(x=1250,y=170,width=150,height=50)


                img  =Image.open(r"image\bellicon.png")
                img = img.resize((40,60),Image.ANTIALIAS)
                self.bellicon = ImageTk.PhotoImage(img)
                Button(cTopFrame,text="rem",command=self.remainderBtn,borderwidth=0,bg="#87CEFA",image=self.bellicon,bd=0,font=('ALGERIAN',14,'bold')).place(x=1430,y=125,width=45,height=70)

                self.rem_label = Label(cTopFrame,text="",textvariable=self.rem,bg="red",fg="white",font=('Algerian',12,"bold"))
                self.rem_label.place(x=1460,y=130)
                
                options = [
                        "Selling ID",
                        "Customer Name",
                        "Mobile No",
                        "City",
                        "Todays",
                        "Total Price",
                        "Remaining Price"
                ]
                
                self.option = StringVar()
                self.option.set("select")
                Label(cTopFrame, text="search by :",font=("Sitka Small Semibold",14),bg="#87CEFA").place(x=720,y=180, height=27)
                dropMenu = OptionMenu(cTopFrame, self.option,*options)
                dropMenu.place(x=850,y=170,width=150,height=40)
                self.option.trace('w',self.search)

                self.c_name.focus()


                if sell_id == None:
                        print("id is null")
                else:
                        print("id is not null")
                        print("total is "+str(total_price))
                        
                        self.total.set(total_price)
                        self.remaining.set(remaining_price)
                        self.sell_id.set(sell_id)
                        

                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                conn = self.create_conn()
                cursor=conn.cursor()

                cursor.execute(""" select date_format(sysdate(),'%y-%m-%d') """)
                date = cursor.fetchall()
                self.getDate=date[0]

                cursor.execute(""" select dayname(sysdate()) """)
                day = cursor.fetchall()
                self.getDay=day[0]

                dateAndDay = str(self.getDay[0])+" / "+str(self.getDate[0])

                conn.close()
                Label(cTopFrame, text="DAY / DATE : ",font=('ALGERIAN',14),bg="#87CEFA").place(x=10,y=180)
                Label(cTopFrame, text=dateAndDay,font=('Sitka Small Semibold',14),bg="#87CEFA",fg="blue").place(x=130,y=180,height=27)
                
                #---------------------------------------------------------------------
                '''------------------------------CALENDAR CODE----------------------'''
                #---------------------------------------------------------------------
                #Button(cTopFrame, text="set remainder",command=self.calendarCode,font=('ALGERIAN',14)).place(x=520,y=175)
                self.calendarCode()
                self.remainder() #this method calculates a count of user thats creadits painding today, and display the no in label
                
                #===================== bind a keys =========================================
                self.root.bind('<Control_L><b>',self.back)
                self.root.bind('<F1>',lambda event : self.c_name.focus())
                self.root.bind('<Control-s>',lambda event : dropMenu.focus())
                self.root.bind('<Control-i>',self.insertBtn)
                self.root.bind('<Control-u>',self.update)
                self.root.bind('<Control-d>',self.delete)
                self.root.bind('<Control-c>',self.clearBtn)
                self.root.bind('<Control-Shift-B>',self.remainderBtn)
                self.root.bind('<Control-r>',self.refresh_treeview)

                '''================================= MAIN FRAME CODE ======================'''
                '''==================================TREEVIEW CODE========================='''
                sc_x = ttk.Scrollbar(treeFrame, orient=HORIZONTAL)
                sc_x.pack(side=BOTTOM, fill=X)

                sc_y = ttk.Scrollbar(treeFrame, orient=VERTICAL)
                sc_y.pack(side=RIGHT, fill=Y)
                
                self.creadit_tree = ttk.Treeview(treeFrame,xscrollcommand=sc_x,yscrollcommand=sc_y)

                self.creadit_tree['columns'] = ("sh_id","s_id","c_name","city","mo_no","total_p","remaining_p","date","rem_date")

                sc_x.config(command=self.creadit_tree.xview)
                sc_y.config(command=self.creadit_tree.yview)

                self.creadit_tree.column("#0", stretch=NO,width=0)
                self.creadit_tree.column("sh_id",minwidth=100,width=50,anchor=CENTER)
                self.creadit_tree.column("s_id",minwidth=100,width=50,anchor=CENTER)
                self.creadit_tree.column("c_name",minwidth=400,anchor=CENTER)
                self.creadit_tree.column("city",minwidth=150,width=150,anchor=CENTER)
                self.creadit_tree.column("mo_no",minwidth=250,anchor=CENTER)
                self.creadit_tree.column("total_p",minwidth=130,anchor=CENTER)
                self.creadit_tree.column("remaining_p",minwidth=130,anchor=CENTER)
                self.creadit_tree.column("date",minwidth=130,anchor=CENTER)
                self.creadit_tree.column("rem_date",minwidth=130, anchor=CENTER)
                
                self.creadit_tree.bind("<ButtonRelease-1>",self.creadit_tree_focus)
                self.creadit_tree.bind("<Up>",self.creadit_tree_up)
                self.creadit_tree.bind("<Down>",self.creadit_tree_down)
                self.creadit_tree.bind("<Enter>",lambda *args : self.creadit_tree.focus(""))
                

                self.creadit_tree.heading("#0", text="")
                self.creadit_tree.heading("sh_id", text="SHOP ID", anchor=CENTER)
                self.creadit_tree.heading("s_id", text="SELLING ID", anchor=CENTER)
                self.creadit_tree.heading("c_name", text="CUSTOMER NAME", anchor=CENTER)
                self.creadit_tree.heading("city", text="CITY", anchor=CENTER)
                self.creadit_tree.heading("mo_no", text="MOBILE NO", anchor=CENTER)
                self.creadit_tree.heading("total_p", text="TOTAL", anchor=CENTER)
                self.creadit_tree.heading("remaining_p", text="CREADIT", anchor=CENTER)
                self.creadit_tree.heading("date", text="DATE", anchor=CENTER)
                self.creadit_tree.heading("rem_date", text="Remainder",anchor=CENTER)

                self.creadit_tree.pack(fill=BOTH,expand=True)

                self.refresh_treeview()

        #=============================sub methods===============================
        def back(self,*args):
                self.root.withdraw()
                self.stock_w = Toplevel()
                obj = stock(self.stock_w)
                
        def create_conn(self):
                conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                return conn

        def calendarCode(self):
                self.sel = StringVar()
                
                def mu_upd(*args):
                        dt = self.sel.get()

                        try:
                                if(len(dt) > 3):
                                    dt1 = datetime.strptime(dt,'%m/%d/%y')
                                    dt2 = dt1+timedelta(days=10)
                                    #self.cal.config(mindate=dt1)
                                    self.cal.config(maxdate=dt2)
            
                        except Exception:
                                messagebox.showerror("","something went wrong plz. reOpen the s/w")

                Label(self.root, text="set remainder :-", font=("Sitka Small Semibold",14),bg="#87CEFA").place(x=380 ,y=187,height=30)              
                #self.cal =DateEntry(self.root, sekectmode="day",state='readonly', font=("Times New Roman",12) ,textvariable=self.sel,mindate=date.today())
                self.cal =DateEntry(self.root, sekectmode="day",state='readonly', font=("Times New Roman",12) ,textvariable=self.sel)
                self.cal.place(x=550,y=187,width=100, height=30)

                self.sel.trace('w',mu_upd)
                
                
                
        def refresh_treeview(self,*args):
                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                conn = self.create_conn()
                cursor=conn.cursor()

                cursor.execute("""select *from creadit WHERE sh_id={}""".format(self.shop_no[0]))
                data = cursor.fetchall()

                self.creadit_tree.delete(*self.creadit_tree.get_children())  #clr treeview

                #print(data)
                conn.close()                

                self.count=0

                for a in data:
                        self.creadit_tree.insert(parent='', index='end', iid=self.count,text="parent",values=(a[0],a[1],a[2],a[3],a[4],a[5],a[6],a[7],a[8]))
                        self.count = self.count+1
                        

        def creadit_tree_focus(self,*args):
                try:
                        cursor_row = self.creadit_tree.focus()
                        content = self.creadit_tree.item(cursor_row)

                        focused_row = content['values']

                        self.name.set(focused_row[2])
                        self.mono.set(focused_row[4])
                        self.total.set(focused_row[5])
                        self.remaining.set(focused_row[6])
                        self.sell_id.set(focused_row[1])
                        self.cit.set(focused_row[3])

                        date_str = focused_row[8]
                        #print(date_str, "date_str")
                        date_object = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()   #convert string to date
                        #print(date_object)
                        date = date_object.strftime("%m/%d/%y")
                        #print(date)  # date to date format
                        self.sel.set(date)

                except Exception:
                        pass

        def creadit_tree_up(self,event=""):
                
                cursor_row = self.creadit_tree.focus()

                try:
                        if int(cursor_row) > 0:
                                content = self.creadit_tree.item(int(cursor_row)-1)
                                focused_row = content['values']

                                self.name.set(focused_row[2])
                                self.mono.set(focused_row[4])
                                self.total.set(focused_row[5])
                                self.remaining.set(focused_row[6])
                                self.sell_id.set(focused_row[1])
                                self.cit.set(focused_row[3])

                                date_str = focused_row[8]
                                #print(date_str, "date_str")
                                date_object = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()   #convert string to date
                                #print(date_object)
                                date = date_object.strftime("%m/%d/%y") #change date format
                                self.sel.set(date)

                except Exception:
                        pass

        def creadit_tree_down(self, event=""):
                cursor_row  = self.creadit_tree.focus()

                try:
                        if int(cursor_row) < int(self.count-1):
                                content = self.creadit_tree.item(int(cursor_row)+1)
                                focused_row = content['values']

                                self.name.set(focused_row[2])
                                self.mono.set(focused_row[4])
                                self.total.set(focused_row[5])
                                self.remaining.set(focused_row[6])
                                self.sell_id.set(focused_row[1])
                                self.cit.set(focused_row[3])

                                date_str = focused_row[8]
                                date_object = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()   #convert string to date
                                date = date_object.strftime("%m/%d/%y") #change the format of date
                                self.sel.set(date)

                except Exception:
                         pass

        #==========================method section===============================

        def insertBtn(self, *args):
                c_name = self.c_name.get()
                c_mono = self.c_mono.get()
                c_city = self.city.get()
                sellId = self.selling_id.get()
                total = self.total_price.get()
                remaining = self.remaining_price.get()

                if c_name=="" or c_mono=="" or c_city=="" or sellId=="" or total=="" or remaining=="":
                        messagebox.showinfo("Info","All fields are required")

                else:
                        try:
                                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                                conn = self.create_conn()
                                cursor=conn.cursor()

                                cursor.execute(""" select count(s_id) from creadit where mo_no = {} AND sh_id={}""".format(c_mono,self.shop_no[0]))
                                ch = cursor.fetchall()

                                #if user ac is clear
                                if ch[0][0]==0:
                                        cursor.execute("""INSERT INTO creadit(sh_id, s_id, c_name, c_city , mo_no, total, creadit, date,rem_date) VALUES({},{},"{}","{}","{}",{},{},"{}","{}" )""".format(self.shop_no[0],sellId,c_name,c_city,c_mono,total,remaining,self.getDate[0],self.cal.get_date()))
                                        conn.commit()

                                                                
                                        #print(self.shop_no[0],sellId,c_name,c_city,c_mono,total,remaining,self.getDate[0])
                                        conn.close()

                                        messagebox.showinfo("Info","Record saved in databse..")

                                #if user have already loan
                                else:
                                    a=messagebox.askquestion("asking","this user already has \"{0}\" creadits do you want to take this another one.".format(ch[0][0]))   
                                    if a=='yes':
                                        conn = self.create_conn()
                                        cursor = conn.cursor()
                                        cursor.execute("""INSERT INTO creadit(sh_id, s_id, c_name, c_city , mo_no, total, creadit, date,rem_date) VALUES({},{},"{}","{}","{}",{},{},"{}","{}" )""".format(self.shop_no[0],sellId,c_name,c_city,c_mono,total,remaining,self.getDate[0],self.cal.get_date()))
                                        conn.commit()

                                        conn.close()

                                        messagebox.showinfo("Info","Record saved in databse..")

                                self.clearBtn()
                                
                        except Exception as e:
##                              messagebox.showerror("NOTE : ","Something went wront plz. check details")
                                print(e)
                self.refresh_treeview()
                self.remainder()

        def update(self, *args):
                try:
                        a=messagebox.askquestion("Note","you want to update this record..")

                        if a=='yes':
                                c_name = self.c_name.get()
                                c_mono = self.c_mono.get()
                                c_city = self.city.get()
                                sellId = self.selling_id.get()
                                total = self.total_price.get()
                                remaining = self.remaining_price.get()
                                
                                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                                conn = self.create_conn()
                                cursor=conn.cursor()

                                cursor.execute("""UPDATE creadit SET c_name='{}', c_city='{}', mo_no='{}',total={},creadit={}, rem_date='{}' WHERE sh_id={} AND s_id={}""".format( c_name, c_city, c_mono, total, remaining,self.cal.get_date(),self.shop_no[0] ,sellId ))
                                conn.commit()
                                
                                conn.close()

                                messagebox.showinfo("Info","Update successfully..")

                except Exception:
                        messagebox.showerror("error","plz select correct selling id..")

                self.clearBtn()
                self.refresh_treeview()
                self.remainder()

        
        def delete(self, *args):
                try:
                        a=messagebox.askquestion("Note","are you sure to delete this record..")

                        if a == 'yes':
                                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                                conn = self.create_conn()
                                cursor=conn.cursor()

                                cursor.execute("DELETE FROM creadit WHERE s_id = {} AND sh_id={}".format(self.selling_id.get(),self.shop_no[0]))
                                conn.commit()
                                
                                conn.close()
                                messagebox.showinfo('info',"deleted successfully..")

                                self.clearBtn()

                        self.refresh_treeview()
                        self.remainder()
                        
                except Exception:
                        messagebox.showerror("info","plz select correct selling ID.")
                
        def clearBtn(self, *args):
                a=messagebox.askquestion("Note","are you sure to clear all textBoxes")

                if a == 'yes':
                        self.city.delete(0,END)
                        self.selling_id.delete(0,END)
                        self.remaining_price.delete(0,END)
                        self.total_price.delete(0,END)
                        self.c_mono.delete(0,END)
                        self.c_name.delete(0,END)
                        self.option.set("select")
                        self.calendarCode()        
                
                self.c_name.focus()
                self.remainder()

        def search(self,*args):
                def setValue(data):
                        self.creadit_tree.delete(*self.creadit_tree.get_children())  #clr treeview               

                        count=0

                        for a in data:
                                self.creadit_tree.insert(parent='', index='end', iid=count,text="parent",values=(a[0],a[1],a[2],a[3],a[4],a[5],a[6],a[7],a[8]))
                                count = count+1

                try:
                        #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                        conn = self.create_conn()
                        cursor=conn.cursor()

                        if self.option.get() == "Selling ID":
                                cursor.execute("SELECT *FROM creadit WHERE s_id = {} AND sh_id = {}".format(self.sell_id.get(),self.shop_no[0]))
                                data = cursor.fetchall()
                                setValue(data)
                                
                                
                        elif self.option.get() == "Customer Name":
                                cursor.execute("SELECT *FROM creadit WHERE c_name = '{}' AND sh_id = {}".format(self.name.get(),self.shop_no[0]))
                                data = cursor.fetchall()
                                setValue(data)

                        elif self.option.get() == "Mobile No":
                                cursor.execute("SELECT *FROM creadit WHERE mo_no = '{}' AND sh_id = {}".format(self.mono.get(),self.shop_no[0]))
                                data = cursor.fetchall()
                                setValue(data)

                        elif self.option.get() == "City":
                                cursor.execute("SELECT *FROM creadit WHERE c_city = '{}' AND sh_id = {}".format(self.cit.get(),self.shop_no[0]))
                                data = cursor.fetchall()
                                setValue(data)

                        elif self.option.get() == "Todays":
                                cursor.execute("SELECT *FROM creadit WHERE date = '{}' AND sh_id = {}".format(self.getDate[0],self.shop_no[0]))
                                data = cursor.fetchall()
                                setValue(data)
                                

                        elif self.option.get() == "Total Price":
                                cursor.execute("SELECT *FROM creadit WHERE total = {} AND sh_id = {}".format(self.total.get(),self.shop_no[0]))
                                data = cursor.fetchall()
                                setValue(data)

                        elif self.option.get() == "Remaining Price":
                                cursor.execute("SELECT *FROM creadit WHERE creadit = {} AND sh_id = {}".format(self.remaining.get(),self.shop_no[0]))
                                data = cursor.fetchall()
                                setValue(data)

                        else:
                                pass
                        conn.close()
                        
                except Exception:
                        messagebox.showerror("Error","Someting went wrong")
                        
                
        def remainderBtn(self, *args):
                try:
                        #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                        conn = self.create_conn()
                        cursor=conn.cursor()

                        data = []
                        count = 0
                        for i in self.rem_ids:
                                cursor.execute("""select *from creadit WHERE sh_id={} AND s_id = {} """.format(self.shop_no[0],i))
                                data.insert(count ,cursor.fetchall())
                                count = count+1
                        #print("MainData == ",data)

                                

                        self.creadit_tree.delete(*self.creadit_tree.get_children())  #clr treeview

                        conn.close()

                        self.creadit_tree.tag_configure('badDebs',background="Red",foreground="white", font=('Times New Roman',12,'bold'))


                        
                        count=0
                        for a in data:
                                if a[0][1] not in self.badDeb:
                                        self.creadit_tree.insert(parent='', index='end', iid=count,text="parent",values=(a[0][0],a[0][1],a[0][2],a[0][3],a[0][4],a[0][5],a[0][6],a[0][7],a[0][8]))
                                        count = count+1
                                if a[0][1] in self.badDeb:
                                        self.creadit_tree.insert(parent='', index='end', iid=count,text="parent",values=(a[0][0],a[0][1],a[0][2],a[0][3],a[0][4],a[0][5],a[0][6],a[0][7],a[0][8]),tags=('badDebs',))
                                        count = count+1
                                        
                except Exception as e:
                        messagebox.showinfo("","something went wrong")
                        print(e)
                        
        def remainder(self):
                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                conn = self.create_conn()
                cursor=conn.cursor()

                rem_ids = []

                cursor.execute("SELECT s_id , rem_date FROM creadit WHERE sh_id = {}".format(self.shop_no[0]))
                self.rem_dates= cursor.fetchall()

                count = 0
                for i in self.rem_dates:
                        #print("rem ids =  = ",i[0]) #print s_id's
                        #print("rem dates = ",i[1]) #print date's
                        
                        strdate = str(i[1])
                        
                        if strdate.find(self.getDate[0]) >= 0:
                                #print("rem_ids = ",i[0])
                                rem_ids.insert(count,i[0])
                                count = count+1

                dcount = 0 #count badDebts
                self.badDeb = [] #store badDebts id's
                for i in self.rem_dates: #remainder_dates calculates in remainder function
                        #print("i[0] = ",i[0]) #s_id
                        #print("i[1] = ",i[1]) #rem_date

                        
                        cursor.execute("SELECT DATEDIFF('{}','{}') ".format(i[1],self.getDate[0] ))
                        temp= cursor.fetchone()

                        #print("date diff of {} is '{}'".format(i,temp[0]))

                        if temp[0] < 0:
                                if i[0] not in rem_ids:
                                        self.badDeb.insert(dcount,i[0])
                                        dcount = dcount+1

                ''' "self.rem_ids" is a combination of creadit id's + bad debtors id's'''
                self.rem_ids = rem_ids+self.badDeb 
                        
                conn.close()
                
                '''set a value of lable that lies on a bell icon'''
                self.rem.set(count+dcount) 



#######################################################################################################################

class Sold:
    def __init__(self,root):
    
        self.root = root
        self.root.state("zoomed")
        self.root.title("Sold Window");
        '''----------------------- shop name and shop no ----------------------------'''
        #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
        conn = self.create_conn()
        cursor = conn.cursor()
        
        cursor.execute("SELECT sh_id FROM shop WHERE sh_status='yes' ")
        self.shop_no = cursor.fetchone()
        cursor.execute("SELECT sh_name FROM shop WHERE sh_id={} ".format(self.shop_no[0]))
        self.shop_name = cursor.fetchone()
        cursor.execute("INSERT INTO sell_id(abc) VALUES('{}')".format('abc'))
        conn.commit()
        cursor.execute("SELECT s_id FROM sell_id order by s_id desc")
        self.sellId = cursor.fetchone()

        conn.close()
        
        print("shop no = "+str(self.shop_no[0]))

        '''-------------------------------FRAMES---------------------------------------'''
        
        #self.mainmid_frame = Label(self.root,image=self.photo).place(x=0,y=0)
        
        sTopFrame = Frame(self.root,relief=RIDGE,bd=10)
        sTopFrame.place(x=0,y=0,width=1530,height=80)

        #sMainFrame = Frame(self.root,bd=10, relief=RIDGE,bg="#53eee5")
        sMainFrame = Frame(self.root,bd=10, relief=RIDGE)
        sMainFrame.place(x=0,y=80,width=1530,height=755)
        sMainFrame.focus()


        img1  =Image.open(r"image\counter_image6.jpg")
        img1 = img1.resize((1508,850),Image.ANTIALIAS)
        self.photo = ImageTk.PhotoImage(img1)
        self.mainmid_frame = Label(sMainFrame,image=self.photo).place(x=0,y=0)

        treeFrame = Frame(sMainFrame, bd=8,relief=RIDGE)
        treeFrame.place(x=18,y=220,width=1000,height=400)

        '''-------------------------------FRAMES---------------------------------------'''

        sTopFramefont=('Times New Roman',22,'bold')
        img  =Image.open(r"image\back.png")
        img  = img.resize((125,50),Image.ANTIALIAS)
        self.backlogo = ImageTk.PhotoImage(img)
        back_btn = Button(sTopFrame,image=self.backlogo,borderwidth=0,command=self.back).place(x=2,y=6)

        
        #----------------- Variables -----------------
        self.getDate = StringVar()
        self.getDay= StringVar()

        
        self.productName = StringVar()
        self.sellingId = StringVar()
        self.qty = IntVar()
        self.price = IntVar()
        shopname = StringVar()
        self.total_price = IntVar()
        self.cashPaid = IntVar()
        self.changeRs = IntVar()

        self.shop_name = StringVar()
        
        self.total_price.set(0)
        #----------------------------set date and day------------
        #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
        conn = self.create_conn()
        cursor=conn.cursor()

        cursor.execute(""" select date_format(sysdate(),'%y-%m-%d') """)
        date = cursor.fetchall()
        self.getDate=date[0]

        cursor.execute(""" select dayname(sysdate()) """)
        day = cursor.fetchall()
        self.getDay=day[0]

        cursor.execute(""" select sh_name from shop where sh_status='yes' """)
        shname = cursor.fetchone()
        self.shop_name=shname[0]
        
        shname_len = len(shname[0])
        self.qty.set(1)
        
        conn.close()
        
        #-------------------------------------------------------------------------------
        Label(sTopFrame,text=self.getDate,font=('Algerian',20,'bold'),fg="#FF26AE").place(x=1100,y=3)
        Label(sTopFrame, text = self.getDay,font=('Algerian',20,'bold'),fg="#09C446").place(x=1300,y=25)
        


        Label(sMainFrame, text="Product Name",bg="#87CEFA", font=('Algerian',18,'bold')).place(x=20,y=10)
        self.p_name = Entry(sMainFrame,textvariable=self.productName, font=('Algerian0',15,'bold'),bd=8,relief=RIDGE)
        self.p_name.place(x=20,y=45,width=250)
        self.p_name.focus()
        
        Label(sMainFrame, text="QTY", font=('Algerian',18,'bold'),bg="#87CEFA").place(x=330,y=10)
        self.p_qty = Entry(sMainFrame,textvariable=self.qty, font=('Algerian0',15,'bold'),bd=8,relief=RIDGE)
        self.p_qty.place(x=330,y=45,width=180)

        Label(sMainFrame, text="PRICE", font=('Algerian',18,'bold'),bg="#87CEFA").place(x=200,y=105)
        self.p_price = Entry(sMainFrame,textvariable=self.price, font=('Algerian0',15,'bold'),bd=8,relief=RIDGE)
        self.p_price.place(x=280,y=100,width=180)


        Label(sMainFrame, text="Total Price",font=('Times New Roman',18),bg="#87CEFA").place(x=1300,y=365)
        self.total_e = Entry(sMainFrame, bd=5,state='disable',textvariable=self.total_price,relief=RIDGE,font=('Times New Roman',18,'bold'))
        self.total_e.place(x=1300,y=400,height=35,width=180)

        Label(sMainFrame, text="Cash Paid RS -",font=('Times New Roman',18),bg="#87CEFA").place(x=1300,y=450)
        self.cashpaid_e = Entry(sMainFrame,textvariable=self.cashPaid ,bd=5,relief=RIDGE,font=('Times New Roman',18,'bold'))
        self.cashpaid_e.place(x=1300,y=480,height=35,width=180)

        Label(sMainFrame, text="Change RS -",font=('Times New Roman',18),bg="#87CEFA").place(x=1300,y=540)
        self.change_e = Entry(sMainFrame, bd=5,textvariable=self.changeRs,state='disable',relief=RIDGE,font=('Times New Roman',18,'bold'))
        self.change_e.place(x=1300,y=570,height=35,width=180)

        self.sellid = Entry(sMainFrame,state='disable',font=('Times New Roman',12,'bold'),textvariable=self.sellingId)
        self.sellid.place(x=20,y=100,width=100)
        self.sellingId.set(self.sellId[0])        

        self.root.bind('<F1>',lambda event="":self.p_name.focus())

        img1  =Image.open(r"image\add.jpg")
        img1 = img1.resize((120,35),Image.ANTIALIAS)
        self.addbtn = ImageTk.PhotoImage(img1)
        Button(sMainFrame,image=self.addbtn,borderwidth=0,command=self.sold_treeview_to_main_treeview).place(x=550,y=45)

        img2  =Image.open(r"image\update.png")
        img2 = img2.resize((120,35),Image.ANTIALIAS)
        self.updatebtn = ImageTk.PhotoImage(img2)
        Button(sMainFrame, image=self.updatebtn,borderwidth=0,command=self.update).place(x=700,y=45)
        
        img3  =Image.open(r"image\delete.png")
        img3 = img3.resize((120,40),Image.ANTIALIAS)
        self.deletebtn = ImageTk.PhotoImage(img3)
        Button(sMainFrame, image=self.deletebtn,command=self.delete).place(x=480,y=90,width=120,height=35)

        img4  =Image.open(r"image\clear.png")
        img4 = img4.resize((120,40),Image.ANTIALIAS)
        self.clrbtn = ImageTk.PhotoImage(img4)
        Button(sMainFrame,image=self.clrbtn,borderwidth=0,command=self.clr).place(x=650,y=90)



        

        if shname_len <= 22:
                Label(sMainFrame, text="WELCOME",font=("Algerian",22,'bold'),fg="#FFFFFF",bg="#4169E1").place(x=1100,y=10)
                l1=Label(sMainFrame, text="TO",font=("Algerian",22,'bold'),fg="#FFFFFF",bg="#4169E1").place(x=1100,y=40)

                Label(sMainFrame, text=self.shop_name,font=("Algerian",22,'bold'),fg="#FFFFFF",bg="#4169E1").place(x=1100,y=70)

        else:
                Label(sMainFrame, text="WELCOME",font=("Algerian",22,'bold'),fg="#FFFFFF",bg="#4169E1").place(x=950,y=10)
                l1=Label(sMainFrame, text="TO",font=("Algerian",22,'bold'),fg="#FFFFFF",bg="#4169E1").place(x=950,y=40)
                Label(sMainFrame, text=self.shop_name,font=("Algerian",22,'bold'),fg="#FFFFFF",bg="#4169E1").place(x=950,y=70)



        sc_x = ttk.Scrollbar(treeFrame, orient=HORIZONTAL)
        sc_x.pack(side=BOTTOM, fill=X)

        sc_y = ttk.Scrollbar(treeFrame, orient=VERTICAL)
        sc_y.pack(side=RIGHT, fill=Y)
        
        self.soldMain_tree = ttk.Treeview(treeFrame,xscrollcommand=sc_x,yscrollcommand=sc_y)

        self.soldMain_tree['columns'] = ("p_id","p_name","p_qty","p_price","t_price")

        sc_x.config(command=self.soldMain_tree.xview)
        sc_y.config(command=self.soldMain_tree.yview)

        self.soldMain_tree.column("#0", stretch=NO,width=0)
        self.soldMain_tree.column("p_id",minwidth=100,width=150,anchor=CENTER)
        self.soldMain_tree.column("p_name",minwidth=400,anchor=CENTER)
        self.soldMain_tree.column("p_qty",minwidth=130,anchor=CENTER)
        self.soldMain_tree.column("p_price",minwidth=130,anchor=CENTER)
        self.soldMain_tree.column("t_price",minwidth=130,anchor=CENTER)
        self.soldMain_tree.bind("<ButtonRelease-1>",self.sell_treeview_focus)

        self.soldMain_tree.heading("#0", text="")
        self.soldMain_tree.heading("p_id", text="PRODUCT ID", anchor=CENTER)
        self.soldMain_tree.heading("p_name", text="PRODUCT NAME", anchor=CENTER)
        self.soldMain_tree.heading("p_qty", text="PRODUCT QTY", anchor=CENTER)
        self.soldMain_tree.heading("p_price", text="PRODUCT PRICE", anchor=CENTER)
        self.soldMain_tree.heading("t_price", text="TOTAL PRICE", anchor=CENTER)

        self.soldMain_tree.pack(fill=BOTH,expand=True)


        

        

        img5  =Image.open(r"image\save2.png")
        img5 = img5.resize((140,50),Image.ANTIALIAS)
        self.insertlogo = ImageTk.PhotoImage(img5)
        Button(sMainFrame,text="SAVE",image=self.insertlogo,font=('Times New Roman',18),bg="#E0C1FF",command=self.save_btn).place(x=950,y=650)

        img6  =Image.open(r"image\print and save.png")
        img6 = img6.resize((180,60),Image.ANTIALIAS)
        self.printAndSavelogo = ImageTk.PhotoImage(img6)
        Button(sMainFrame,image=self.printAndSavelogo,borderwidth=0,command=self.printAndSave_btn).place(x=1100,y=650)

        #----------------------------------------------------------------------------------------------
        self.productmainframeOpen = False
        self.productName.trace('w',self.pro_nameEnter)
        self.cashPaid.trace('w',self.cash_paid)
        
        self.count_sell_treeviewValues=0
        self.count_down_cursor_value=0
        self.highlightrow= None

        self.target = -1

        
        #____________________________________________________________________________________________
        #_______________________________ BIND EVENTS ON MAIN FRAME __________________________________
        
        self.root.bind('<Control_L><b>',self.back)
        self.p_name.bind('<Down>',self.p_nameDown)
        self.p_name.bind('<Up>',self.p_nameUp)
        self.p_name.bind('<Return>',self.p_nameEnter)
        
        self.p_qty.bind('<Return>',self.sold_treeview_to_main_treeview)
        self.p_price.bind('<Return>',self.sold_treeview_to_main_treeview)

        #local btn binding
        '''PRODUCT NAME ENTRY BIND'''
        self.p_name.bind('<Control_L> <a>',self.sold_treeview_to_main_treeview),self.p_name.bind('<Control_L> <u>',self.update)
        self.p_name.bind('<Control_L> <d>',self.delete),self.p_name.bind('<Control_L> <c>',self.clr)
        '''PRODUCT QTY ENTRY BIND'''
        self.p_qty.bind('<Control_L> <a>',self.sold_treeview_to_main_treeview),self.p_qty.bind('<Control_L> <u>',self.update)
        self.p_qty.bind('<Control_L> <d>',self.delete),self.p_qty.bind('<Control_L> <c>',self.clr)
        '''PRODUCT PRICE ENTRY BIND'''
        self.p_price.bind('<Control_L> <a>',self.sold_treeview_to_main_treeview),self.p_price.bind('<Control_L> <u>',self.update)
        self.p_price.bind('<Control_L> <d>',self.delete),self.p_price.bind('<Control_L> <c>',self.clr)
        '''PRODUCT CASH PAID ENTRY BIND'''
        self.cashpaid_e.bind('<Control_L> <a>',self.sold_treeview_to_main_treeview),self.cashpaid_e.bind('<Control_L> <u>',self.update)
        self.cashpaid_e.bind('<Control_L> <d>',self.delete),self.cashpaid_e.bind('<Control_L> <c>',self.clr)
        self.cashpaid_e.bind('<Control_L> <s>',self.save_btn),self.cashpaid_e.bind('<Control_L> <p>',self.printAndSave_btn)

        #_______________________________________________________________________________________________
        #_____________________________________ METHOD SECTION __________________________________________

    def create_conn(self):
        conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
        return conn

    def update(self,*args):
        if self.p_name.get()=="" or self.p_qty.get()=="" or self.p_price.get()=="":
                messagebox.showinfo("Message","All Fields are required..")

        else:
                try:
                        total_price =float(self.p_qty.get()) * float(self.p_price.get())
                        
                        #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                        conn = self.create_conn()
                        cursor=conn.cursor()

                        
                        cursor.execute(" UPDATE selling_table SET qty='{}',s_price='{}',total_price='{}' WHERE p_id='{}' AND sh_id='{}' AND s_id='{}' ".format(
                                self.p_qty.get(), self.p_price.get() , total_price,self.main_focus_row[0], self.shop_no[0],self.sellid.get() 
                            ))
                        conn.commit()
                        conn.close()
                        self.refresh_mainTreeview_data()

                except Exception:
                        messagebox.showinfo("Info","Plz select correct product")
                
    def delete(self,*args):
        if self.p_name.get()=="" or self.p_qty.get()=="" or self.p_price.get()=="":
                messagebox.showinfo("Message","All Fields are required..")
        else:
                try:
                        #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                        conn = self.create_conn()
                        cursor=conn.cursor()

                        cursor.execute(" DELETE FROM selling_table WHERE p_id='{}' AND sh_id='{}' AND s_id='{}' ".format(
                                self.main_focus_row[0],self.shop_no[0],self.sellid.get() 
                            ))
                        conn.commit()
                        conn.close()
                        self.refresh_mainTreeview_data()
                        self.changeRs.set(0)
                        
                except Exception:
                        messagebox.showinfo("Info","Plz select correct product")

    def clr(self,*args):
        self.p_name.delete(0,END)
        self.qty.set(1)
        self.p_price.delete(0,END)
        self.p_name.focus()
        
        
    def back(self,*args):
        self.root.withdraw()
        self.stock_w = Toplevel()
        obj = stock(self.stock_w)

      
    def sell_treeview_focus(self,*args):
        self.clr()
        cursor_row = self.soldMain_tree.focus()
        content = self.soldMain_tree.item(cursor_row)
        self.main_focus_row = content['values']
        try:
                self.productName.set(self.main_focus_row[1])
                self.qty.set(self.main_focus_row[2])
                self.price.set(self.main_focus_row[3])
        except Exception:
                self.productName.set("")
                self.qty.set(1)
                self.price.set(0)

    def cash_paid(self,*args):
        if self.total_e.get() == "" or self.total_e.get() == None:
                pass

        elif self.total_e.get()=='0':
                messagebox.showinfo("Message","Plz Select Product First..")
                self.cashPaid.set(0)

        else:
                try:
                        cash_paid = float(self.cashpaid_e.get())
                        total = float(self.total_e.get())
                        
                        change =  cash_paid -  total
                        self.changeRs.set(change)
                except Exception:
                        if self.cashpaid_e.get() == "":
                                self.changeRs.set(0)
                        elif self.cashpaid_e.get()=='+' or self.cashpaid_e.get()=='-':
                                if self.cashpaid_e.get()=='-':
                                        self.cashPaid.set('-')
                                else:
                                        self.cashPaid.set('+')
                        else:
                                self.cashPaid.set(0)
                                self.changeRs.set(0)
                                messagebox.showinfo("Message","Plz Enter Number Values Only..")
                    
    def go_to_CreaditPage(self,sell_id, total_price, remaining_price):
        print("sell id = ",sell_id," total price = ",total_price, " remaining price = ",remaining_price)
        self.creadit_w = Toplevel(self.root)
        obj = Creadit(self.creadit_w,sell_id, total_price, remaining_price)
        
       
    
    def save_btn(self,*args):
        if self.total_e.get() == "" or self.total_e.get() == None or self.total_e.get()=='0':
                messagebox.showinfo("Message","Plz Select Product First..")
                self.p_name.focus()

        elif self.cashpaid_e.get() == '0'or self.cashpaid_e.get() == None or self.cashpaid_e.get() == "":
                messagebox.showinfo("Message","Plz receive payment first..")
                self.cashpaid_e.focus()
                return

        elif float(self.change_e.get()) >= 0:
                #here the code...
                #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                conn = self.create_conn()
                cursor=conn.cursor()


                cursor.execute(""" SELECT p_id, qty FROM selling_table WHERE sh_id='{}' AND s_id={}""".format(self.shop_no[0],self.sellId[0]))
                sell_data = cursor.fetchall()

                cursor.execute(""" SELECT count(p_id) FROM selling_table WHERE sh_id={} AND s_id={}""".format(self.shop_no[0],self.sellId[0]))
                count = cursor.fetchone()

                cursor.execute(""" SELECT p_id,p_qty FROM products WHERE sh_id={} AND p_id in (SELECT p_id FROM selling_table WHERE sh_id={} AND s_id={})""".format(
                        self.shop_no[0],self.shop_no[0],self.sellId[0],
                    ))
                product_data = cursor.fetchall()

                #print("count = ",count[0])

                for i in range(count[0]):
                        #print("===================================")
                        sell_value = sell_data[i][1]
                        product_value = product_data[i][1]
                        pro_id = sell_data[i][0]
                        #print("product id = ",pro_id)
                        #print("product data = ",product_value)
                        #print("sell data = ",sell_value)
                        sub =  int(product_value) - int(sell_value)
                        #print(sub)


                        cursor.execute(""" UPDATE products SET p_qty={} WHERE sh_id={} AND p_id={} """.format(sub,self.shop_no[0],pro_id))
                        conn.commit()

                cursor.execute(" INSERT INTO confirm_selling(sh_id,s_id,s_date) VALUES({},{},'{}') ".format(self.shop_no[0],self.sellId[0],self.getDate[0]))
                conn.commit()

                
                messagebox.showinfo("message","saved")
                conn.close()

                
        else:
                #transfer the sellid into creadit page
                ch = messagebox.askquestion("question","this payment is -ve do you want to go to this payment in creadit")
                if ch == 'yes':
                        conn = self.create_conn()
                        cursor=conn.cursor()


                        cursor.execute(""" SELECT p_id, qty FROM selling_table WHERE sh_id='{}' AND s_id={}""".format(self.shop_no[0],self.sellId[0]))
                        sell_data = cursor.fetchall()

                        cursor.execute(""" SELECT count(p_id) FROM selling_table WHERE sh_id={} AND s_id={}""".format(self.shop_no[0],self.sellId[0]))
                        count = cursor.fetchone()

                        cursor.execute(""" SELECT p_id,p_qty FROM products WHERE sh_id={} AND p_id in (SELECT p_id FROM selling_table WHERE sh_id={} AND s_id={})""".format(
                                self.shop_no[0],self.shop_no[0],self.sellId[0],
                            ))
                        product_data = cursor.fetchall()

                        #print("count = ",count[0])

                        for i in range(count[0]):
                                #print("===================================")
                                sell_value = sell_data[i][1]
                                product_value = product_data[i][1]
                                pro_id = sell_data[i][0]
                                #print("product id = ",pro_id)
                                #print("product data = ",product_value)
                                #print("sell data = ",sell_value)
                                sub =  int(product_value) - int(sell_value)

        ##                        ------------

                                
                                sell_id=self.sellId[0]
                                total_price=self.total_e
                                remaining_price=self.change_e

                                conn = self.create_conn()
                                cursor=conn.cursor()

                                cursor.execute(""" UPDATE products SET p_qty={} WHERE sh_id={} AND p_id={} """.format(sub,self.shop_no[0],pro_id))
                                conn.commit()
                                
                                cursor.execute(""" INSERT INTO confirm_selling(sh_id,s_id,s_date) VALUES({},{},'{}') """.format(self.shop_no[0],sell_id,self.getDate[0]))
                                conn.commit()

                                conn.close()

                                #total price, remaining price and selling id transfer into creadit page
                                self.go_to_CreaditPage(sell_id, self.total_e.get(), float(self.total_e.get()) - float(self.cashpaid_e.get()))
                else:
                        print('saved click')                        
        self.p_name.focus()

        #clr the treeview or entry boxes
        self.soldMain_tree.delete(*self.soldMain_tree.get_children())  #delete treeview
        self.cashPaid.set(0)
        self.changeRs.set(0)
        self.total_price.set(0)

        #find new sell id
        conn = self.create_conn()
        cursor=conn.cursor()
        
        cursor.execute("INSERT INTO sell_id(abc) VALUES('{}')".format('abc'))
        conn.commit()
        cursor.execute("SELECT s_id FROM sell_id order by s_id desc")
        self.sellId = cursor.fetchone()

        conn.close()

        self.sellingId.set(self.sellId[0])

        
        #after clicking save btn, delete unwanted selling
        conn = self.create_conn()
        cursor=conn.cursor()
        
        cursor.execute("""delete from sell_id where s_id not in(select s_id from confirm_selling)""")
        conn.commit()
        
        cursor.execute("""delete from selling_table where s_id not in (select s_id from confirm_selling)""");
        conn.commit()
        conn.close()


    def printAndSave_btn(self,*args):
        print("print and save")

    #=======================================================
    def pro_nameEnter(self,*args):
        if self.p_name.get()== "":
                self.productmainframeOpen = False
                self.target=-1
                self.productmainframe.destroy()
    
            
        elif self.productmainframeOpen == False:
                self.productmainframe = LabelFrame(self.root,text="select product",font=('Algerian',12,'bold'),bd=10,relief=RIDGE)
                self.productmainframe.place(x=100,y=310,width=800,height=400)

                self.productFrame = Frame(self.productmainframe,bd=5,relief=RIDGE)
                self.productFrame.place(x=3,y=35,width=780,height=335)

                #-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-
                self.count_down_cursor_value = 0
                #-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-_-
                
                img  =Image.open(r"image\exit.png")
                img = img.resize((90,35),Image.ANTIALIAS)
                self.exit = ImageTk.PhotoImage(img)
                Button(self.productmainframe, image=self.exit,command=self.exitbtn).place(x=690,y=0)


                sc_x = ttk.Scrollbar(self.productFrame, orient=HORIZONTAL)
                sc_x.pack(side=BOTTOM, fill=X)

                sc_y = ttk.Scrollbar(self.productFrame, orient=VERTICAL)
                sc_y.pack(side=RIGHT, fill=Y)

                self.sold_tree = ttk.Treeview(self.productFrame,xscrollcommand=sc_x,yscrollcommand=sc_y)
                self.sold_tree['columns'] = ("p_id","p_name","p_qty","p_price")

                self.sold_tree.column("#0",stretch=NO,width=0)
                self.sold_tree.column("p_id",minwidth=100,width=100,anchor=CENTER)
                self.sold_tree.column("p_name",minwidth=400,anchor=CENTER)
                self.sold_tree.column("p_qty",minwidth=100,anchor=CENTER)
                self.sold_tree.column("p_price",minwidth=150,anchor=CENTER)

                self.sold_tree.heading("#0",text="")
                self.sold_tree.heading("p_id",text="product id",anchor=CENTER)
                self.sold_tree.heading("p_name",text="product name",anchor=CENTER)
                self.sold_tree.heading("p_qty",text="QTY",anchor=CENTER)
                self.sold_tree.heading("p_price",text="price",anchor=CENTER)
                self.sold_tree.pack(fill=BOTH, expand=True)

                #self.sold_tree_insert(-1)
                self.sold_tree_insert()

                self.productmainframeOpen = True
            
        else:
                #self.sold_tree_insert(0)
                self.sold_tree_insert()
            
            

    #----------------------------product frame functions------------------------------
    def exitbtn(self):
            self.productmainframeOpen = False
            self.target=-1 #new line add
            self.productmainframe.destroy()

    def sold_tree_insert(self):
        
            proName = '%'+self.p_name.get()+'%'

            #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
            conn = self.create_conn()
            cursor=conn.cursor()

            cursor.execute("SELECT p_id,p_name,p_qty,s_price FROM products WHERE p_name like '{}' AND sh_id = '{}'  ".format(proName,self.shop_no[0]))
            self.soldTree_data = cursor.fetchall()

            cursor.execute(" SELECT count(p_name)as count FROM products WHERE p_name like '{}' AND sh_id = '{}' ".format(proName,self.shop_no[0]))
            c= cursor.fetchone()
            self.count_sell_treeviewValues = c[0]        
            conn.close()

            #print(self.count_sell_treeviewValues)
            
            self.sold_tree.delete(*self.sold_tree.get_children())  #clr treeview

            self.sold_tree.tag_configure('focus',background="blue",foreground="white", font=('Times New Roman',12,'bold'))
            self.sold_tree_insert_refresh()
        
                
    def sold_tree_insert_refresh(self):
            count = 0
            for i in self.soldTree_data:
               
                    if self.target == count:
                            self.sold_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3]),tags=('focus',))

                            self.content =self.sold_tree.item(self.target)
                            
                            self.highlightrow = self.content['values']
                            
                    else:
                            self.sold_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3]))
                    count = count + 1
        

    def p_nameUp(self,*args):
            if self.p_name.get() == "":
                    pass
            
            elif self.target != 0 and self.target != -1:
                    self.target = self.target-1
                    self.sold_tree_insert()
            
            else:
                    self.target = self.count_sell_treeviewValues-1
                    self.sold_tree_insert()
            
    
    def p_nameDown(self,event=""):
            if self.p_name.get() == "":
                    pass

            elif self.target < self.count_sell_treeviewValues-1:
                    self.target = self.target+1
                    self.sold_tree_insert()
            else:
                    self.target = 0
                    self.sold_tree_insert()
            

    def p_nameEnter(self,*args):
            if self.p_name.get()=="" or self.highlightrow == None or self.target == -1:
                    pass

            else:
                    
                    '''data fiill in the text boxes'''
                    p_name = self.highlightrow[1] 
                    p_price = self.highlightrow[3]
                        
                    self.price.set(p_price)
                    self.productName.set(p_name)
                    self.exitbtn()
             


    def sold_treeview_to_main_treeview(self,*args):
            if self.p_name.get()=="":
                    messagebox.showinfo("message","product name is required.")
            else:
                    try:
                            sh_id = self.shop_no[0]
                            s_id  = self.sellId[0]
                            p_name = self.highlightrow[1]
                            p_qty = self.p_qty.get()
                            #p_price = p_price[0]
                            s_price = self.p_price.get()
                            total_price = float(s_price) * float(p_qty)
                            date = self.getDate[0]
                            p_id = self.highlightrow[0]
                            
                    
                            #_______________________________________________
                            #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                            conn = self.create_conn()
                            cursor=conn.cursor()
                            cursor.execute(" SELECT p_price,p_qty FROM products WHERE p_id='{}' AND sh_id='{}' ".format(p_id,sh_id))
                            price_qty = cursor.fetchone()
                            conn.close()
                            
                            #------------------------------------------------

                            p_price = price_qty[0]
                            main_pro_qty = price_qty[1]
                            #________________________________________________
                            #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                            conn = self.create_conn()
                            cursor=conn.cursor()
                    except Exception:
                        messagebox.showinfo("message","select correct product")
                        return

                    
                    try:
                            if int(p_qty) < int(main_pro_qty):
                                    cursor.execute(" insert into selling_table values({},{},'{}',{},{},{},{},'{}',{})".format(
                                            sh_id,s_id,p_name,p_qty,p_price,s_price,total_price,date,p_id
                                        ))
                                    conn.commit()
                                    conn.close()
                            else:
                                    messagebox.showinfo("message","qty not available")
                                    return
                        
                    except Exception:
                            messagebox.showinfo("message","product already available")
                    self.refresh_mainTreeview_data()

    def refresh_mainTreeview_data(self):
            sh_id = self.shop_no[0]
            s_id  = self.sellId[0]
            
            #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
            conn = self.create_conn()
            cursor=conn.cursor()     
            cursor=conn.cursor()
            cursor.execute(" SELECT p_id,p_name,qty,s_price,total_price FROM selling_table WHERE s_id='{}' AND sh_id='{}' ".format(
                    s_id,sh_id
                ))
            data = cursor.fetchall()
            
            conn.close()
            #------------------------------------------------
            self.soldMain_tree.delete(*self.soldMain_tree.get_children()) #clear treeview
            count = 0
            for i in data:
                    self.soldMain_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3],i[4]))
                    count = count + 1

            self.clr()

            #conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
            conn = self.create_conn()
            cursor=conn.cursor()
            cursor.execute(" SELECT SUM(total_price) FROM selling_table WHERE sh_id='{}' AND s_id='{}' ".format(sh_id,s_id))
            all_total_price = cursor.fetchone()
            conn.close()

            if all_total_price[0] == None:
                    self.total_price.set(0)
            else:
                    self.total_price.set(all_total_price[0])

            self.cash_paid()
        


#######################################################################################################################

class firstPage:
    def __init__(self,root):
        self.root = root
        #self.root.geometry("1550x1000+0+0")
        #self.root.overrideredirect(True) #hide the title bar
        
        self.root.state("zoomed")
        self.root.resizable(False,False)
        self.root.title("Login Window");

        #___________________________________________________________
        topFrame = LabelFrame(self.root,relief=RIDGE,bd=5 ,text="select your choice")
        topFrame.place(x=5,y=10,width=630,height=100)

        '''midFrame = Frame(self.root,bd=10,relief=RIDGE)
        midFrame.place(x=10, y=100, width=1300,height=750)'''
        #-----------------------------------------------------------
        img  =Image.open(r"image\register.png")
        img = img.resize((160,60),Image.ANTIALIAS)
        self.registerlogo = ImageTk.PhotoImage(img)
        Button(topFrame,image=self.registerlogo,text="register",command=self.registerWindow).place(x=10,y=5,width=160,height=60);

        img1  =Image.open(r"image\login.png")
        img1 = img1.resize((160,60),Image.ANTIALIAS)
        self.loginlogo = ImageTk.PhotoImage(img1)
        Button(topFrame,image=self.loginlogo, text="Login window" ,command=self.loginWindow).place(x=220,y=5,width=160,height=60);

        img2  =Image.open(r"image\exit.png")
        img2 = img2.resize((160,60),Image.ANTIALIAS)
        self.exitlogo = ImageTk.PhotoImage(img2)
        Button(topFrame,image=self.exitlogo ,text="Quit" ,command=self.quit).place(x=430,y=5,width=160,height=60);

        #__________________________________________________________
        '''============== register window variables=============='''
        self.register_window_open=False
        self.login_window_open=False
    
     #____________________________________________________________________________________
    '''============================ main window methods ==============================='''
    def create_conn(self):
                conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
                return conn

    def createS_conn(self):
                conn = mysql.connector.connect(host='localhost', database='datasetting', user='root', password='Tushar()mysql[123]')
                return conn
    
    def registerWindow(self):
        if self.login_window_open == True:
            self.loginFrame.destroy()
            self.login_window_open=False
            
        if self.register_window_open == False:
            self.registerFrame = Frame(self.root,bd=10,bg="#53eee5",relief=RIDGE)
            self.registerFrame.place(x=20, y=200, width=900,height=500)

            '''bg  =Image.open(r"image\background.png")
            bg = bg.resize((900,500),Image.ANTIALIAS)
            self.photo = ImageTk.PhotoImage(bg)
            Label(self.registerFrame,image=self.photo).place(x=0,y=0)

            canvas = Canvas(self.registerFrame, width= 900, height=500)
            canvas.pack()
            bg  =Image.open(r"image\background.png")
            bg = bg.resize((900,500),Image.ANTIALIAS)
            self.photo = ImageTk.PhotoImage(bg)
            Label(canvas,image=self.photo).place(x=0,y=0)'''

            #canvas.create_text(250,10,text="REGISTER YOUR SHOP",font=('time',18,'bold'))
            reg = Label(self.registerFrame,bg="#53eee5",text="REGISTER YOUR SHOP" , font=("Modern No. 20",22,"bold"))
            reg.place(x=250,y=10)

            Label(self.registerFrame,bg="#53eee5" ,text="YOUR SHOP SH ID : ").place(x=600,y=23)
            ids = Entry(self.registerFrame, bd=5,state='readonly')
            ids.place(x=720,y=23)
            
            Label(self.registerFrame,bg="#53eee5", text="ENTER YOUR SHOP NAME : ").place(x=100,y=80)
            self.e1 = Entry(self.registerFrame, font=('time',14,'bold'),bd=5)
            self.e1.place(x=100,y=110)

            Label(self.registerFrame,bg="#53eee5", text="ENTER Mo.No : ").place(x=500,y=80)
            self.e2 = Entry(self.registerFrame, font=('time',14,'bold'),bd=5)
            self.e2.place(x=500,y=110)
            
            Label(self.registerFrame,bg="#53eee5", text="ENTER PASSWORD : ").place(x=100, y=200)
            self.e3 = Entry(self.registerFrame, font=('time',14,'bold'),show='*',bd=5)
            self.e3.place(x=100,y=230)
            
            Label(self.registerFrame,bg="#53eee5", text="CONFIRM PASSWORD : ").place(x=500, y=200)
            self.e4 = Entry(self.registerFrame, font=('time',14,'bold'),show='*',bd=5)
            self.e4.place(x=500,y=230)

            self.e1.focus()

            img  =Image.open(r"image\register1.png")
            img = img.resize((290,70),Image.ANTIALIAS)
            self.reglogo = ImageTk.PhotoImage(img)
            Button(self.registerFrame,image=self.reglogo,text="REGISTER NOW",command=self.registerbtn, bd=5).place(x=100,y=300,width=290,height=70)

            img1  =Image.open(r"image\clear1.png")
            img1 = img1.resize((290,70),Image.ANTIALIAS)
            self.clrlogo = ImageTk.PhotoImage(img1)
            Button(self.registerFrame,image=self.clrlogo ,text="CLEAR", bd=5, command=self.clear).place(x=442,y=300,width=290,height=70)

            self.register_window_open=True

        else:
            self.registerFrame.destroy()
            self.register_window_open=False
        

    def loginWindow(self):
        if self.register_window_open == True:
            self.registerFrame.destroy()
            self.register_window_open=False
            
            
        if self.login_window_open == False:
            self.loginFrame = Frame(self.root,bd=10,bg="#53eee5",relief=RIDGE)
            self.loginFrame.place(x=20, y=200, width=900,height=500)

            reg = Label(self.loginFrame,bg="#53eee5", text="LOGIN YOUR SHOP" , font=("Modern No. 20",22,"bold"))
            reg.place(x=280,y=10)

            Label(self.loginFrame,bg="#53eee5", text="ENTER YOUR SHOP NAME : ").place(x=100,y=80)
            self.e1 = Entry(self.loginFrame, font=('time',14,'bold'),bd=5)
            self.e1.place(x=100,y=110)

            Label(self.loginFrame,bg="#53eee5", text="ENTER Mo.No : ").place(x=500,y=80)
            self.e2 = Entry(self.loginFrame, font=('time',14,'bold'),bd=5)
            self.e2.place(x=500,y=110)
            
            Label(self.loginFrame,bg="#53eee5", text="ENTER PASSWORD : ").place(x=100, y=200)
            self.e3 = Entry(self.loginFrame, font=('time',14,'bold'),show='*',bd=5)
            self.e3.place(x=100,y=230)

            self.e1.focus()

            img=Image.open(r"image\login.png")
            img = img.resize((160,40),Image.ANTIALIAS)
            self.loglogo = ImageTk.PhotoImage(img)
            Button(self.loginFrame,image=self.loglogo,text="LOGIN",command=self.loginbtn, bd=5).place(x=400,y=230,width=160,height=40)

            img1=Image.open(r"image\clear2.png")
            img1 = img1.resize((160,40),Image.ANTIALIAS)
            self.cllogo = ImageTk.PhotoImage(img1)
            Button(self.loginFrame,image=self.cllogo , text="CLEAR", bd=5, command=self.clearlogin).place(x=600,y=230,width=160,height=40)
            Button(self.loginFrame, text="forgot password",bg="#53eee5",width=25,fg='red',font=('time',12,'bold'),command=self.forgotbtn,relief='flat').place(x=600,y=300)


            self.login_window_open=True

        else:
            self.loginFrame.destroy()
            self.login_window_open=False

    def quit(self):
        ch = messagebox.askquestion("warning","ARE YOU SURE TO CLOSE THIS APPLICATION")

        if ch=='yes':
            self.root.destroy()

     #____________________________________________________________________________________
    '''============================ register window functions ========================='''
    def registerbtn(self):
        if self.e1.get()=="" or self.e2.get()=="" or self.e3.get()=="" or self.e4.get()=="":
            messagebox.showinfo('info','All fields required..')

        else:
            shop_name = self.e1.get()
            mo_no= self.e2.get()
            
            pas = self.e3.get()
            c_pass = self.e4.get()


            if(pas == c_pass):
                    conn = self.create_conn()
                    cursor = conn.cursor()
                    cursor.execute('''select *from shop WHERE sh_name='{}' and mono='{}' '''.format(shop_name,mo_no))
                    ch = cursor.fetchall()
                    conn.close()
                
                
                    if ch:    
                        messagebox.showinfo('info',"this shop is already register")

                    else:   #if shop is does not register
                         try:
                            conn = self.create_conn()
                            cursor = conn.cursor()
                            cursor.execute(''' update shop set sh_status='no' ''')
                            conn.commit()
                            cursor.execute("INSERT INTO SHOP(sh_name,sh_status,mono,pass) VALUES('{}','{}','{}','{}')".format(shop_name,'yes',mo_no,pas))
                            conn.commit()
                            conn.close()
                         except Exception:
                                messagebox.showinfo('info','mo.no length must be 10')
                            
                         conn = self.create_conn()
                         cursor = conn.cursor()
                         cursor.execute(''' select sh_id from shop where sh_status='yes' ''')
                         self.shop_no = cursor.fetchone()
                         conn.close()

                         self.defaultSetting_orderQty()
                         messagebox.showinfo("info","Register successfully")

                         self.root.withdraw()
                         self.stock_w = Toplevel()
                         obj = stock(self.stock_w)
                         print("go to stock window..")
                       
            else:
                messagebox.showinfo("info","password and confirm password must be same")
                


    def clear(self):
        self.e1.delete(0,END)
        self.e2.delete(0,END)
        self.e3.delete(0,END)
        self.e4.delete(0,END)
        self.e1.focus()

     #____________________________________________________________________________________
    '''============================ login window functions ========================='''
    def loginbtn(self):
        if self.e1.get()=="" or self.e2.get()=="" or self.e3.get()=="":
            messagebox.showinfo('info','all fields are required');

        else:
            shop_name = self.e1.get()
            mo_no = self.e2.get()
            pas = self.e3.get()

            print(shop_name,mo_no,pas)
            conn = self.create_conn()
            cursor = conn.cursor()
            cursor.execute('''select *from shop WHERE sh_name='{}' and mono='{}' and pass='{}' '''.format(shop_name,mo_no,pas))
            ch = cursor.fetchall()
            conn.close()

            if ch:
                conn = self.create_conn()
                cursor = conn.cursor()

                cursor.execute('''UPDATE shop SET sh_status='no' WHERE sh_status='yes' ''')
                conn.commit()
                
                cursor.execute('''update shop set sh_status='yes' WHERE sh_name='{}' and mono='{}' and pass='{}' '''.format(shop_name,mo_no,pas))
                conn.commit()

                conn.close()

                
##                messagebox.showinfo('info','go to stock page');
                self.root.withdraw()
                self.stock_w = Toplevel()
                obj = stock(self.stock_w)
                print("go to stock page")
                
            else:
                messagebox.showinfo("info","this shop is unavailable");
            
            
    def forgotbtn(self):
        pass
    
    def clearlogin(self):
        self.e1.delete(0,END)
        self.e2.delete(0,END)
        self.e3.delete(0,END)
        self.e1.focus()

    def defaultSetting_orderQty(self): #this function sets the default order value for order page this is useful for auto order..
        sconn = self.createS_conn()
        cursor = sconn.cursor()
        cursor.execute("INSERT INTO order_ws(sh_id,l_stock_qty,m_stock_qty,saving_path) VALUES({},{},{},'{}')".format(self.shop_no[0],10,20,f"C:/shop management/order_recipts/"))
        sconn.commit()
        sconn.close()


#######################################################################################################################
class stock:
    def __init__(self,root):
        self.root = root

        self.root = root
        self.root.state("zoomed")
        self.root.resizable(False,False)
        self.root.title("Stock page")
        #self.root.geometry("1550x1000")
        #self.root.resizable(False,False) #hide the title bar

        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT sh_id FROM shop WHERE sh_status='yes' ")
        self.shop_no = cursor.fetchone()
        conn.close()
        #------------------- STOCK FRAME VARIABLES -----------
        self.p_id = IntVar()
        self.p_name = StringVar()
        self.sqty = IntVar()
        self.sp_price = IntVar()
        self.ss_price = IntVar()
            
        #------------------- ADD VARIABLES -------------------
        self.ap_name = StringVar()

        
        '''
            ===============================================================================================
        '''
        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT sh_name FROM shop WHERE sh_status='yes' ")
        shop_name = cursor.fetchone()
        conn.close()
        
        sh_name = shop_name[0]
        '''
            ===============================================================================================
        '''

        img1  =Image.open(r"image\bg.png")
        img1 = img1.resize((1600,850),Image.ANTIALIAS)
        self.photo = ImageTk.PhotoImage(img1)
        
        self.mainmid_frame = Label(self.root,image=self.photo).place(x=0,y=0)
        
        topframe=Frame(self.root,bg="lightyellow",bd=10,relief=RIDGE)
        topframe.place(x=0,y=0,width=1533,height=100)
        
        

        btnframe = Frame(self.root,bg="red", bd=10,relief=RIDGE)
        btnframe.place(x=0,y=100,width=150,height=445)

        img  =Image.open(r"image\TE_logo.png")
        img = img.resize((80,80),Image.ANTIALIAS)
        self.photoimg1 = ImageTk.PhotoImage(img)
        Button(topframe,text="click",command=self.logo,image=self.photoimg1,borderwidth=0).place(x=0,y=0)
        
        Label(topframe, text=sh_name,bg="lightyellow",fg="red",font=("Algerian",30)).pack(side=TOP , fill=X,padx=100,pady=25)

        img2  =Image.open(r"image\home1.png")
        img2 = img2.resize((125,50),Image.ANTIALIAS)
        self.homelogo = ImageTk.PhotoImage(img2)
        Button(btnframe,image=self.homelogo ,text="HOME",command=self.home,font=('Sitka Small Semibold',12,'bold'),bd=5,relief=GROOVE).place(x=2,y=3,height=50,width=125)

        img3  =Image.open(r"image\stock.png")
        img3 = img3.resize((125,50),Image.ANTIALIAS)
        self.stocklogo = ImageTk.PhotoImage(img3)
        Button(btnframe,image=self.stocklogo ,text="STOCK",command=self.stock,font=('Sitka Small Semibold',12,'bold'),bd=5,relief=GROOVE).place(x=2,y=55,height=50,width=125)
        
        img4  =Image.open(r"image\sold.png")
        img4 = img4.resize((125,50),Image.ANTIALIAS)
        self.soldlogo = ImageTk.PhotoImage(img4)
        Button(btnframe,image=self.soldlogo, text="SOLD", command=self.sold,font=('Sitka Small Semibold',12,'bold'),bd=5,relief=GROOVE).place(x=2,y=108,height=50,width=125)

        img5  =Image.open(r"image\creadit.png")
        img5 = img5.resize((125,50),Image.ANTIALIAS)
        self.creaditlogo = ImageTk.PhotoImage(img5)
        Button(btnframe,image=self.creaditlogo, text="CREDIT", command=self.creadit,font=('Sitka Small Semibold',12,'bold'),bd=5,relief=GROOVE).place(x=2,y=160,height=50,width=125)

        img6  =Image.open(r"image\order goods.png")
        img6 = img6.resize((125,50),Image.ANTIALIAS)
        self.orderlogo = ImageTk.PhotoImage(img6)
        Button(btnframe,image=self.orderlogo, text="ORDER GOODS",command=self.order,font=('Sitka Small Semibold',11,'bold'),bd=5,relief=GROOVE).place(x=2,y=213,height=50,width=125)

        img7  =Image.open(r"image\accounting.png")
        img7 = img7.resize((125,50),Image.ANTIALIAS)
        self.acclogo = ImageTk.PhotoImage(img7)
        Button(btnframe,image=self.acclogo,text="ACCOUNTING", command=self.accounting,font=('Sitka Small Semibold',12,'bold'),bd=5,relief=GROOVE).place(x=2,y=266, height=50,width=125)

        img8  =Image.open(r"image\notes.png")
        img8 = img8.resize((125,50),Image.ANTIALIAS)
        self.noteslogo = ImageTk.PhotoImage(img8)
        Button(btnframe,image=self.noteslogo, text="NOTES",command=self.notes,font=('Sitka Small Semibold',12,'bold'),bd=5,relief=GROOVE).place(x=2,y=319,height=50,width=125)

        img9  =Image.open(r"image\log out.png")
        img9 = img9.resize((125,50),Image.ANTIALIAS)
        self.logoutlogo = ImageTk.PhotoImage(img9)
        Button(btnframe,image=self.logoutlogo,text="LOG OUT", command=self.log_out,font=('Sitka Small Semibold',12,'bold'),bd=5,relief=GROOVE).place(x=2,y=372, height=50,width=125)


        '''========================BIND GLOBAL ROOT BUTTONS=========================='''
        self.root.focus()

        self.root.bind('<Control-Shift-H>',self.home)
        self.root.bind('<Control-Shift-S>',self.stock)
        self.root.bind('<F2>',self.sold)
        self.root.bind('<Control-Shift-C>',self.creadit)
        self.root.bind('<Control-Shift-O>',self.order)
        self.root.bind('<Control-Shift-A>',self.accounting)
        self.root.bind('<Control-Shift-N>',self.notes)
        self.root.bind('<Control-Shift-L>',self.log_out)

        '''=========================================================================='''
            
        self.stock_w_open = False
        self.searchCount = 0

    def create_conn(self):
        conn = mysql.connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
        return conn
        
    def logo(self):
        pass

    def home(self,*args):
        if self.stock_w_open == True:
            self.stock_frame.destroy()
            self.stock_w_open = False

    def stock(self,*args):
        if self.stock_w_open == False:
            self.stock_frame = Frame(self.root,bg='white',bd=10,relief=RIDGE)
            self.stock_frame.place(x=250,y=100,width=920,height=500) 
            

            Label(self.stock_frame, text="STOCK AVAILABLE ON SHOP",font=("Sitka Small Semibold",15,"bold")).pack(side=TOP,fill=X)

            self.o_frame = Frame(self.stock_frame,bd=10,relief=RIDGE)
            self.o_frame.place(x=0,y=30,width=900,height=250)

            self.stock_btm_frame = Frame(self.stock_frame,bg='red',bd=10,relief=RIDGE)
            self.stock_btm_frame.place(x=0,y=280,width=900,height=200)

            stock_left_frame = Frame(self.stock_btm_frame, bd=10,relief=RIDGE)
            stock_left_frame.place(x=1,y=0,width=660,height=180)
       
        
            stock_right_frame = Frame(self.stock_btm_frame, bd=10,relief=RIDGE)
            stock_right_frame.place(x=665,y=0,width=220,height=180)
            #-------------------------------------------------------------------------------------------
            sc_x = ttk.Scrollbar(self.o_frame, orient=HORIZONTAL)
            sc_x.pack(side=BOTTOM, fill=X)

            sc_y = ttk.Scrollbar(self.o_frame, orient=VERTICAL)
            sc_y.pack(side=RIGHT, fill=Y)
            
            self.order_tree = ttk.Treeview(self.o_frame,xscrollcommand=sc_x,yscrollcommand=sc_y)

            self.order_tree['columns'] = ("p_id","p_name","p_qty","p_price","sold_price")

            sc_x.config(command=self.order_tree.xview)
            sc_y.config(command=self.order_tree.yview)

            self.order_tree.column("#0", stretch=NO,width=0)
            self.order_tree.column("p_id",anchor=CENTER, minwidth=100,width=100)
            self.order_tree.column("p_name",minwidth=250,anchor=CENTER)
            self.order_tree.column("p_qty",minwidth=130,anchor=CENTER)
            self.order_tree.column("p_price",minwidth=130,anchor=CENTER)
            self.order_tree.column("sold_price",minwidth=130,anchor=CENTER)
            self.order_tree.column("#0", stretch=NO,width=0)
            self.order_tree.bind("<ButtonRelease-1>",self.stock_cursor)

            self.order_tree.heading("#0", text="")
            self.order_tree.heading("p_id", text="PRODUCT ID", anchor=CENTER)
            self.order_tree.heading("p_name", text="PRODUCT NAME", anchor=CENTER)
            self.order_tree.heading("p_qty", text="PRODUCT QTY", anchor=CENTER)
            self.order_tree.heading("p_price", text="PRODUCT PRICE", anchor=CENTER)
            self.order_tree.heading("sold_price", text="SELLING PRICE", anchor=CENTER)
            self.order_tree.heading("#0", text="")
           
            self.clr_treeview()

            self.order_tree.pack(fill=BOTH)
            #--------------------------------------------------------------------------------------------
            Label(stock_left_frame, text="P NAME", font=('time',12,'bold')).place(x=65,y=3)
            self.product = Entry(stock_left_frame,textvariable=self.p_name, font=('time',12,'bold'),bd=5)
            self.product.place(x=10,y=40,width=180)
            
            Label(stock_left_frame, text="QTY", font=('time',12,'bold')).place(x=210,y=3)
            self.qty = Entry(stock_left_frame,textvariable=self.sqty, font=('time',12,'bold'),bd=5)
            self.qty.place(x=210,y=40,width=120)
            
            Label(stock_left_frame, text="P PRICE", font=('time',12,'bold')).place(x=350,y=3)
            self.p_price = Entry(stock_left_frame,textvariable=self.sp_price, font=('time',12,'bold'),bd=5)
            self.p_price.place(x=350,y=40,width=140)

            Label(stock_left_frame, text="S PRICE", font=('time',12,'bold')).place(x=520,y=3)
            self.s_price = Entry(stock_left_frame,textvariable=self.ss_price, font=('time',12,'bold'),bd=5)
            self.s_price.place(x=500,y=40,width=140)
            self.stock_clere()

            img  =Image.open(r"image\insert.png")
            img = img.resize((140,30),Image.ANTIALIAS)
            self.insertlogo = ImageTk.PhotoImage(img)
            insert_btn = Button(stock_left_frame, text="INSERT",image = self.insertlogo,borderwidth=0,command=self.stock_insert)
            insert_btn.place(x=10, y=90,width=130)

            
            img  =Image.open(r"image\update.png")
            img = img.resize((140,30),Image.ANTIALIAS)
            self.updatelogo = ImageTk.PhotoImage(img)
            Button(stock_left_frame, text="UPDATE",image=self.updatelogo,borderwidth=0, command=self.stock_update).place(x=170, y=90,width=130)

            img  =Image.open(r"image\delete.png")
            img = img.resize((140,30),Image.ANTIALIAS)
            self.deletelogo = ImageTk.PhotoImage(img)
            Button(stock_left_frame, text="DELETE",borderwidth=0,image=self.deletelogo ,command=self.stock_delete).place(x=330, y=90,width=130)

            img  =Image.open(r"image\clear.png")
            img = img.resize((140,30),Image.ANTIALIAS)
            self.clrlogo = ImageTk.PhotoImage(img)
            Button(stock_left_frame, text="CLERE",image=self.clrlogo, borderwidth=0 ,command=self.stock_clere).place(x=480, y=90,width=150)

            img  =Image.open(r"image\reseat.png")
            img = img.resize((140,30),Image.ANTIALIAS)
            self.reseatlogo = ImageTk.PhotoImage(img)
            Button(stock_left_frame, text="REFRESH",image= self.reseatlogo,borderwidth=0,command=self.refresh).place(x=490, y=130,width=130)


            img  =Image.open(r"image\search.png")
            img = img.resize((140,30),Image.ANTIALIAS)
            self.searchlogo = ImageTk.PhotoImage(img)
            Button(stock_left_frame, text="SEARCH",image= self.searchlogo,borderwidth=0,command=self.search).place(x=330, y=130,width=130)
   

            img  =Image.open(r"image\high stock.png")
            img = img.resize((175,70),Image.ANTIALIAS)
            self.hstocklogo = ImageTk.PhotoImage(img)
            Button(stock_right_frame, text="HIGH STOCK",image=self.hstocklogo,borderwidth=0,command=self.highStockFun).place(x=15, y=8, width=175,height=70)

            img  =Image.open(r"image\btn.png")
            img = img.resize((175,70),Image.ANTIALIAS)
            self.lstocklogo = ImageTk.PhotoImage(img)
            Button(stock_right_frame, text="LOW STOCK",image=self.lstocklogo,borderwidth=0,command=self.lowStockFun).place(x=15, y=85, width=175,height=70)

            
            
            self.stock_w_open = True

            '''===================================== BINDING BUTTONS============================='''
            
            
            #local button binds
            '''EntryBoxList = [self.product,self.qty,self.p_price,self.s_price]
            for i in EntryBoxList:
                i.bind('<Control_L><i>',self.stock_insert)
                i.bind('<Control_L><u>',self.stock_update)
                i.bind('<Control_L><d>',self.stock_delete)
                i.bind('<Control_L><c>',self.stock_clere)
                i.bind('<Control_L><r>',self.refresh)
                i.bind('<Control_L><s>',self.search)
                i.bind('<Control_L><h>',self.highStockFun)
                i.bind('<Control_L><l>',self.lowStockFun)'''

            if self.stock_w_open == True:
                try:
                     self.root.bind('<Control_L><i>',self.stock_insert)
                     self.root.bind('<Control_L><u>',self.stock_update)
                     self.root.bind('<Control_L><d>',self.stock_delete)
                     self.root.bind('<Control_L><c>',self.stock_clere)
                     self.root.bind('<Control_L><r>',self.refresh)
                     self.root.bind('<Control_L><s>',self.search)
                     self.root.bind('<Control_L><h>',self.highStockFun)
                     self.root.bind('<Control_L><l>',self.lowStockFun)

                except Exception:
                    pass
                    
            '''=================================================================================='''


    def sold(self,*args):
        self.root.withdraw()
        self.sold_w = Toplevel()
        obj = Sold(self.sold_w)
        print("sold class object hear")

    def creadit(self,*args):
        self.root.withdraw()
        self.creadit_w = Toplevel()
        obj = Creadit(self.creadit_w,None,None,None)
        print("creadit class object hear")

    def order(self,*args):
        self.root.withdraw()
        self.order_w = Toplevel()
        obj = Order(self.order_w)        
        print("order class object hear")

    def accounting(self,*args):
        self.root.withdraw()
        self.accounting_w = Toplevel()
        obj = Accounting(self.accounting_w)        
        print("accounting class object hear")

    def notes(self,*args):
        print("notes class object hear")

    def log_out(self,*args):
        self.c=messagebox.askquestion("warning","ARE YOU SURE TO LOG OUT YOUR SHOP AC.")
        if self.c == "yes":
            messagebox.showinfo("info","LOG OUT SUCCESSFULLY")

            conn = self.create_conn()
            cursor = conn.cursor()
            cursor.execute("UPDATE shop set sh_status='no' WHERE sh_status='yes'")
            conn.commit()
            conn.close()

            self.root.withdraw()
            self.Register_w = Toplevel()
            obj = firstPage(self.Register_w)
            print("log out successfully")

        else:
            #nothing
            pass

    #___________________________________________________________________________________________
    '''=====================================stock w functions ================================'''
    def stock_insert(self,*args):
        try:
            if self.product.get()=='' or self.qty.get()=='' or self.p_price.get()=='' or self.s_price.get()=='':
                messagebox.showerror("error", "all fields are required")

            else:
                conn = self.create_conn()
                cursor = conn.cursor()
              
                sh_id=self.shop_no[0]
                
                cursor.execute("select *from products where p_name ='{}' ".format(self.p_name.get()))
                avl = cursor.fetchone()

                if avl:
                    messagebox.showerror('error', "product already available")

                else:
                    cursor.execute("""INSERT INTO products(sh_id, p_name, p_qty, p_price, s_price) VALUES('{}','{}','{}','{}','{}')""".format
                                (sh_id,self.p_name.get(),self.qty.get(),self.p_price.get(), self.s_price.get() ))
                    conn.commit()
                    conn.close()
                    self.stock_clere()
                    #messagebox.showinfo('INFO','insert successfully')
                    self.clr_treeview()

        except Exception:
            pass

    def stock_update(self,*args):
        try:
            if self.product.get()=='' or self.qty.get()=='' or self.p_price.get()=='' or self.s_price.get()=='':
                messagebox.showerror("error", "all fields are required")

            else:
                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("UPDATE products set p_name='{}', p_qty='{}', p_price='{}', s_price='{}' where p_id ='{}' ".format(
                            self.p_name.get(),self.qty.get(),self.p_price.get(), self.s_price.get(),self.p_id
                    ))            
                conn.commit()
                conn.close()
                self.stock_clere()
                #messagebox.showinfo('INFO','update successfully')
                self.clr_treeview()
                
        except Exception:
            pass

    def stock_delete(self,*args):
        try:
            if self.product.get()=='' or self.qty.get()=='' or self.p_price.get()=='' or self.s_price.get()=='':
                messagebox.showerror("error", "all fields are required")

            else:
                self.ch = messagebox.askquestion("warning","ARE YOU SURE TO DELETE THIS RECORD...")
                if self.ch == "yes":
                    conn = self.create_conn()
                    cursor = conn.cursor()
                    cursor.execute("DELETE from products WHERE p_id='{}'".format(self.p_id))
                    conn.commit()
                    conn.close()
                    self.stock_clere()
                    self.clr_treeview()
                    
        except Exception:
            pass

    def refresh(self,*args):
        try:
            self.clr_treeview()
        except Exception:
            pass
                
    def stock_clere(self,*args):
        try:
            self.product.focus()
            self.product.delete(0,END)
            self.qty.delete(0,END)
            self.p_price.delete(0 ,END)
            self.s_price.delete(0,END)
        except Exception:
            pass

    def search(self,*args):
        try:
            if self.searchCount == 0:
                self.stock_search_frame = Frame(self.root, bd=10,relief=RIDGE)
                self.stock_search_frame.place(x=600,y=600,width=300,height=180)
                self.searchCount=1

                options = [
                    "By ID",
                    "By Name"
                ];
                self.option = StringVar()
                self.option.set("Search By")

                dropMenu = OptionMenu(self.stock_search_frame, self.option, *options)
                dropMenu.place(x=3,y=3)

                self.option.trace('w',self.search_OptionMenu_Click)
                
                #Button(self.stock_search_frame, text="SEARCH",image= self.rsearchlogo,borderwidth=0,command=self.itemSearch).place(x=200, y=40,width=40)

            else:
                self.stock_search_frame.destroy()
                self.searchCount=0
        except Exception:
            pass

    def highStockFun(self,*args):
        try:
            conn = self.create_conn()
            cursor = conn.cursor()
            
            sh_id = self.shop_no[0]
            
            cursor.execute("SELECT p_id,p_name,p_qty,p_price,s_price FROM products where sh_id='{}' and p_qty>='{}' order by p_qty desc".format(sh_id,18))
            data = cursor.fetchall()
            conn.close()
            
            self.order_tree.delete(*self.order_tree.get_children())  #clere treeview

            self.order_tree.tag_configure('highstock',background="lime", font=('Times New Roman',12,'bold'))

            count = 0
            for i in data:
                self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3],i[4]),tags=('highstock',))
                count += 1
        except Exception:
            pass
        
    def lowStockFun(self,*args):
        try:
            conn = self.create_conn()
            cursor = conn.cursor()
           
            sh_id = self.shop_no[0]
            cursor.execute("SELECT p_id,p_name,p_qty,p_price,s_price FROM products where sh_id='{}' and p_qty<='{}' order by p_qty asc".format(sh_id,10))
            data = cursor.fetchall()
            conn.close()
            
            self.order_tree.delete(*self.order_tree.get_children())  #clere treeview

            self.order_tree.tag_configure('lowstock',background="DarkOrange", font=('Times New Roman',12,'bold'))

            count = 0
            for i in data:
                self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3],i[4]),tags=('lowstock',))
                count += 1
                
        except Exception:
            pass
    #___________________________________________________________________________________________

    def clr_treeview(self):
        self.order_tree.delete(*self.order_tree.get_children())  #clere treeview

        conn = self.create_conn()
        cursor=conn.cursor()
        
        sh_id = self.shop_no
        cursor.execute("SELECT p_id,p_name,p_qty,p_price,s_price FROM products where sh_id='{}' order by p_qty asc".format(sh_id[0]))
        data = cursor.fetchall()
        conn.close()

        self.order_tree.tag_configure('lowstock',background="DarkOrange", font=('Times New Roman',12,'bold'))
        self.order_tree.tag_configure('midstock',background="yellow", font=('Times New Roman',12,'bold'))
        self.order_tree.tag_configure('highstock',background="lime", font=('Times New Roman',12,'bold'))
        
        count = 0
        for i in data:
            if i[2]<=10:
                self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3],i[4]),tags=('lowstock',))
                count += 1
            if i[2]<=20 and i[2]>=11:
                self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3],i[4]),tags='midstock',)
                count += 1
            if i[2]>=21:
                self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[0],i[1],i[2],i[3],i[4]),tags='highstock',)
                count += 1
        
    
    def stock_cursor(self, event=""):
        try:
            cursor_row = self.order_tree.focus()
            content = self.order_tree.item(cursor_row)
            row = content['values']
            self.p_id = row[0]
            self.p_name.set(row[1])
            self.ss_price.set(row[4])
            self.sp_price.set(row[3])
            self.sqty.set(row[2])
            
        except Exception:
            pass
        self.product.focus()
        
    
    def search_OptionMenu_Click(self,*args):  #this func. returns the value on option menu
        self.str_out = StringVar()
        self.str_out.set(self.option.get())
        
        self.e1_str = StringVar()
        self.searchEntry = Entry(self.stock_search_frame,bd=2,textvariable=self.e1_str,font=('time',10,'bold'),relief=RIDGE)
        self.searchEntry.place(x=3,y=40,height=30,width=190)

        img  =Image.open(r"image\round search.png")
        img = img.resize((40,30),Image.ANTIALIAS)
        self.rsearchlogo = ImageTk.PhotoImage(img)
        Button(self.stock_search_frame, text="SEARCH",image= self.rsearchlogo,borderwidth=0,command=self.itemSearch).place(x=200, y=40,width=40)

        '''==================================AUTO FILL LISTBOX CODE================================='''
        conn = self.create_conn()
        cursor = conn.cursor()
        cursor.execute("""SELECT p_name FROM products WHERE sh_id='{}' """.format(self.shop_no[0]))
        d = cursor.fetchall()
        conn.close()
        self.my_searchList=[r for r , in d]

        l1= Listbox(self.stock_search_frame,height=6,relief='flat',bg='SystemButtonFace',highlightcolor='SystemButtonFace')
        l1.place(x=3,y=70,height=80,width=190)

        def my_update(my_widget):
            my_w = my_widget.widget
            index=int(my_w.curselection()[0])
            value=my_w.get(index)
            self.e1_str.set(value)
            l1.delete(0,END)

        def my_down(my_widget):
            l1.focus()
            l1.selection_set(0)

        def get_data(*args):
            search_str = self.searchEntry.get() #user enter string

            if search_str != "":
                l1.delete(0,END)
                for element in self.my_searchList:
                    if(re.search(search_str, element, re.IGNORECASE)):
                        l1.insert(END, element)
            else:
                l1.delete(0,END)

        self.searchEntry.bind('<Down>',my_down)
        l1.bind('<Right>',my_update)
        l1.bind('<Return>',my_update)  #enter btn
        self.e1_str.trace('w',get_data)
        

        
    def itemSearch(self,):
        if self.option.get() == "Search By" or self.searchEntry.get() == "":
            messagebox.showerror("Error","all fields are required");
            

        else:
            check_selection = self.str_out.get()
            entry_value = '%'+self.searchEntry.get()+'%'
            
            
            if check_selection == 'By ID':   #if the user intrested to search a product by its id 
                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("""SELECT *FROM products WHERE p_id like '{}'  AND sh_id ='{}' """.format(entry_value,self.shop_no[0]))
                data = cursor.fetchall()
                conn.close()

                
                
            else:     #if the user intrested to search a product by its id
                conn = self.create_conn()
                cursor = conn.cursor()
                cursor.execute("""SELECT *FROM products WHERE p_name like '{}' AND sh_id='{}' ORDER BY p_qty asc """.format(entry_value,self.shop_no[0]))
                data = cursor.fetchall()
                conn.close()
                
                
            if data:      #check the product is available or not                
                
                self.order_tree.delete(*self.order_tree.get_children())  #clere treeview

                self.order_tree.tag_configure('lowstock',background="DarkOrange", font=('Times New Roman',12,'bold'))
                self.order_tree.tag_configure('midstock',background="yellow", font=('Times New Roman',12,'bold'))
                self.order_tree.tag_configure('highstock',background="lime", font=('Times New Roman',12,'bold'))
                
                count = 0
                for i in data:
                    if i[3]<=10:
                        self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[1],i[2],i[3],i[4],i[5]),tags=('lowstock',))
                        count += 1
                    if i[3]<=17 and i[3]>=11:
                        self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[1],i[2],i[3],i[4],i[5]),tags='midstock',)
                        count += 1
                    if i[3]>=18:
                        self.order_tree.insert(parent="", index="end", iid=count, text="" , values=(i[1],i[2],i[3],i[4],i[5]),tags='highstock',)
                        count += 1

        

            else:           #if the product is not available in our database
                messagebox.showinfo('INFO','This Product Is Unavailable')
                self.order_tree.delete(*self.order_tree.get_children())  #clere treeview'''
    



##################################################################################################################
class splash:
    def __init__(self,root):
        self.root = root
        self.root.geometry("1550x1000")
        self.root.overrideredirect(True) #hide the title bar

        img1  =Image.open(r"image\colorful_bg.jpg")
        img1 = img1.resize((1530,860),Image.ANTIALIAS)
        self.photo = ImageTk.PhotoImage(img1)
        Label(self.root,image=self.photo).place(x=0,y=0)


        self.root.configure(bg='silver')

        Label(self.root, text="STOCK MANAGEMENT SYSTEM", font=("Algerian",30),fg="red").pack(pady=150)

        Label(self.root, text="Developed By : ", font=("Algerian",18)).place(x=1000,y=600)
        Label(self.root, text="TUSHAR MAHAJAN ", font=("Algerian",18)).place(x=1200,y=650)

        
        conn = connector.connect(host='localhost', database='shopmanagement', user='root', password='Tushar()mysql[123]')
        cursor=conn.cursor()

        cursor.execute("SELECT *FROM shop WHERE sh_status='yes' ")
        self.status_ch = cursor.fetchall()
        conn.close()

        self.root.after(2000,self.main_window)
          
    def main_window(self,*args):
     
         self.root.destroy() #destroy a splash window
         #--------------------------

       
         #--------------------------    
         root2 = Tk()   # create a stock window
         
         if self.status_ch:   # if any shop are logged in move on register window
             print("any shop is register");
             obj = stock(root2)

             #go to stock page
            
         else:                      # if no any shop are logged in move on stock page
             print("no any shop are register")
             #go to 1st main page
             obj = firstPage(root2)
         #---------------------------
        

if __name__ == '__main__':
    root = Tk()
    obj = splash(root)
    root.mainloop()



